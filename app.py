"""
Lab Planning & Occupancy Dashboard — Streamlit Web App
=======================================================
All 3 tools in one web interface.
Run: streamlit run app.py
"""

import os
import sys
import math
import tempfile
import io
import datetime

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import lab_core as core
import generators as gen_module
from generators import merge_files, GROUP_A, GROUP_B, GROUP_C, ALL_TYPES

# ──────────────────────────────────────────────────────────
#  PAGE CONFIG
# ──────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Lab Planning Dashboard",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────
#  CUSTOM CSS
# ──────────────────────────────────────────────────────────
st.markdown("""
<style>
    /* Main background */
    .stApp { background-color: #F0F4F8; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1F3864 0%, #2E75B6 100%);
    }
    section[data-testid="stSidebar"] * { color: white !important; }
    section[data-testid="stSidebar"] .stRadio label { color: white !important; }
    section[data-testid="stSidebar"] p { color: #BDD7EE !important; font-size: 13px; }

    /* ── Chat input: override the sidebar wildcard rule above ─────────────── */
    section[data-testid="stSidebar"] .stTextInput input,
    section[data-testid="stSidebar"] .stTextInput > div > div > input,
    section[data-testid="stSidebar"] input[type="text"] {
        background-color: #FFFFFF !important;
        color: #1F3864 !important;
        caret-color: #1F3864 !important;
        border: 1.5px solid #6FA3D8 !important;
        border-radius: 8px !important;
        font-size: 13px !important;
    }
    section[data-testid="stSidebar"] .stTextInput input::placeholder,
    section[data-testid="stSidebar"] input[type="text"]::placeholder {
        color: #8AAAD4 !important;
        opacity: 1 !important;
    }
    section[data-testid="stSidebar"] .stTextInput input:focus,
    section[data-testid="stSidebar"] input[type="text"]:focus {
        border-color: #5BA3E0 !important;
        box-shadow: 0 0 0 2px rgba(91,163,224,0.3) !important;
        outline: none !important;
    }

    /* Tool headers */
    .tool-header {
        background: linear-gradient(90deg, #1F3864 0%, #2E75B6 100%);
        color: white; padding: 18px 24px; border-radius: 10px;
        margin-bottom: 20px;
    }
    .tool-header h2 { color: white !important; margin: 0; font-size: 22px; }
    .tool-header p  { color: #BDD7EE !important; margin: 4px 0 0 0; font-size: 13px; }

    /* Metric cards */
    .metric-card {
        background: white; border-radius: 8px; padding: 16px 20px;
        border-left: 4px solid #2E75B6; margin: 6px 0;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08);
    }
    .metric-card.red   { border-left-color: #FF4444; }
    .metric-card.green { border-left-color: #70AD47; }
    .metric-card.amber { border-left-color: #FFD700; }

    /* Section headers */
    .section-label {
        font-weight: 700; font-size: 14px; color: #1F3864;
        margin: 20px 0 8px 0; padding-bottom: 4px;
        border-bottom: 2px solid #2E75B6;
    }

    /* Capacity input group */
    .cap-group {
        background: white; border-radius: 8px; padding: 14px 16px;
        border: 1px solid #D9E1F2; margin: 6px 0;
    }

    /* Status badges */
    .badge-red    { background:#FF4444; color:white; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:700; }
    .badge-yellow { background:#FFD700; color:#333; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:700; }
    .badge-green  { background:#70AD47; color:white; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:700; }

    /* Download button */
    .stDownloadButton > button {
        background: #1F5C1A !important; color: white !important;
        border-radius: 8px !important; font-weight: 700 !important;
        padding: 10px 24px !important; font-size: 15px !important;
        width: 100% !important;
    }

    /* Generate button */
    .stButton > button {
        background: #1F3864 !important; color: white !important;
        border-radius: 8px !important; font-weight: 700 !important;
        padding: 10px 24px !important; font-size: 15px !important;
        width: 100% !important;
    }

    /* File uploader */
    [data-testid="stFileUploader"] {
        border: 2px dashed #2E75B6 !important;
        border-radius: 10px !important; padding: 10px !important;
    }

    /* Tab styling */
    .stTabs [data-baseweb="tab-list"] { gap: 4px; }
    .stTabs [data-baseweb="tab"] {
        border-radius: 6px 6px 0 0 !important;
        background: #D9E1F2 !important; color: #1F3864 !important;
        font-weight: 600 !important; padding: 8px 20px !important;
    }
    .stTabs [aria-selected="true"] {
        background: #1F3864 !important; color: white !important;
    }

    div[data-testid="stExpander"] {
        border: 1px solid #D9E1F2 !important;
        border-radius: 8px !important;
    }

    /* Chat bubbles */
    .chat-user {
        background: #1F3864; color: white;
        padding: 8px 12px; border-radius: 12px 12px 2px 12px;
        margin: 4px 0; font-size: 12px; max-width: 92%;
        margin-left: auto; text-align: right;
    }
    .chat-bot {
        background: #E8F0FE; color: #1F3864;
        padding: 8px 12px; border-radius: 12px 12px 12px 2px;
        margin: 4px 0; font-size: 12px; max-width: 92%;
        border-left: 3px solid #2E75B6;
    }
    .chat-wrap { max-height: 320px; overflow-y: auto; padding: 4px; }

    /* Annual summary table */
    .summary-table { width:100%; border-collapse:collapse; font-family:Arial; font-size:13px; margin-top:8px; }
    .summary-table th { background:#1F3864; color:white; padding:10px 12px; text-align:left; font-weight:700; }
    .summary-table td { padding:8px 12px; border-bottom:1px solid #E8EEF7; }
    .summary-table tr:nth-child(even) td { background:#F6F9FF; }
    .summary-table tr:hover td { background:#EEF4FF; }
    .util-red   { background:#FF4444 !important; color:white; font-weight:700; text-align:center; border-radius:4px; }
    .util-amber { background:#FFA500 !important; color:white; font-weight:700; text-align:center; border-radius:4px; }
    .util-green { background:#70AD47 !important; color:white; font-weight:700; text-align:center; border-radius:4px; }
</style>
""", unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────
#  CONSTANTS
# ──────────────────────────────────────────────────────────
THEME_COLORS = {
    "Default (Red/Yellow/Green)": {"high":"#FF4444","medium":"#FFD700","low":"#70AD47"},
    "Blue Gradient":              {"high":"#1F4E79","medium":"#2E75B6","low":"#BDD7EE"},
    "Pastel":                     {"high":"#FF9999","medium":"#FFEB99","low":"#C6EFCE"},
}

CURRENT_YEAR = datetime.date.today().year
CURRENT_WEEK = datetime.date.today().isocalendar()[1]

GROUP_A = {"name":"Mechanical Labs",  "types":["LCF","Creep"],
           "color":"#1A4E8A","cap_total":72}
GROUP_B = {"name":"Coating Labs",     "types":["Cold Spray","HVOF","Plasma"],
           "color":"#1F5C1A","cap_per_lab":350}  # each lab independently 350/yr


# ──────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────

def load_uploaded_files(uploaded_files, allowed_types=None):
    """Save uploaded Streamlit files to temp dir, return paths."""
    paths = []
    for f in uploaded_files:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.write(f.getbuffer())
        tmp.close()
        paths.append(tmp.name)
    return paths


def _persist_upload(cache_key, file_obj):
    """
    Cache an uploaded file's bytes in session_state so it survives
    widget-triggered re-runs (Streamlit re-runs clear file uploader state).
    Returns dict {"bytes": b"...", "name": "..."} or None.
    """
    if file_obj is not None:
        st.session_state[cache_key] = {
            "bytes": file_obj.getbuffer().tobytes(),
            "name":  file_obj.name,
        }
    return st.session_state.get(cache_key)


def _persist_upload_multi(cache_key, file_list):
    """
    Same as _persist_upload but for multiple files (Tool 4 comparison).
    Returns list of {"bytes", "name"} dicts or [].
    """
    if file_list:
        st.session_state[cache_key] = [
            {"bytes": f.getbuffer().tobytes(), "name": f.name}
            for f in file_list
        ]
    return st.session_state.get(cache_key, [])



def _show_file_diagnostics(path, expected_types):
    """Show what's in the file and what was expected — helps users debug wrong-file uploads."""
    import traceback as _tb
    try:
        xl = pd.ExcelFile(path)
        sheets = xl.sheet_names
        raw = pd.read_excel(xl, sheet_name=sheets[0], header=None, nrows=8)
        with st.expander("🔍 File diagnostics — click to see raw file content", expanded=True):
            st.markdown(f"**Sheets found:** {', '.join(sheets)}")
            st.markdown(f"**Expected lab types:** `{expected_types}`")
            st.markdown("**First 8 rows of data:**")
            st.dataframe(raw, use_container_width=True)
            st.markdown("**💡 Tip:** Each tool needs specific lab types in a `Type` column:")
            st.markdown("- 🔵 Tool 1 → `LCF`, `Creep`")
            st.markdown("- 🟢 Tool 2 → `Cold Spray`, `HVOF`, `Plasma`")
            st.markdown("- 🟠 Tool 3 → `Thermal Rig`")
    except Exception as e:
        st.warning(f"Could not read file for diagnostics: {e}")

def _bytes_to_tmp(b: bytes, original_name: str = None) -> str:
    """
    Write bytes to a temp file and return the path.
    If original_name is given, the temp file is named to preserve it
    (so filename-based heuristics like monthly-block type inference work).
    """
    if original_name:
        safe_name = os.path.basename(original_name)
        tmp_dir = tempfile.mkdtemp()
        path = os.path.join(tmp_dir, safe_name)
        with open(path, "wb") as f:
            f.write(b)
        return path
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(b)
    tmp.close()
    return tmp.name


def _detect_and_maybe_select_type(path, allowed_types, tool_key):
    """
    If the file is a monthly-block format (year-named sheets, TOTAL SAMPLES REMOVED)
    but has multiple allowed_types, show a selectbox so the user can tell us
    which lab type this file represents.

    Returns the allowed_types list to pass into load_and_filter — either the
    original list (if format is not monthly-block, or if only one type) or a
    single-element list chosen by the user.
    """
    import lab_core as _core
    try:
        is_block, _, _ = _core._is_monthly_block_format(path)
    except Exception:
        is_block = False

    if not is_block or len(allowed_types) == 1:
        return allowed_types   # no disambiguation needed

    # Multiple types + monthly-block format → ask user
    key = f"_type_sel_{tool_key}"
    st.info(
        "📋 This file uses the **monthly-block format** (year-named sheets with "
        "TOTAL SAMPLES REMOVED rows). Please select which lab type this file contains:"
    )
    chosen = st.selectbox(
        "Which lab type does this file represent?",
        options=allowed_types,
        key=key,
    )
    return [chosen]


def _load_multi_files(paths, allowed_types, tool_key):
    """
    Load and merge 1+ uploaded files for a single tool, restricted to that
    tool's own allowed_types ('only look at the assigned task for this
    tool', even if a file happens to contain other lab types' rows).

    Each file is independently run through the full smart parser
    (core.load_and_filter) — long format, wide format, messy/junk headers,
    data on any sheet, or the monthly-block format (year-named sheets with
    TOTAL SAMPLES REMOVED rows). This means any of those layouts can be
    mixed across files for the same tool.

    Monthly-block files carry no Type column, so if a tool has more than
    one allowed type (e.g. Tool 1 = LCF + Creep) and the filename doesn't
    give away which type a given file represents, the user is asked once
    per file via a selectbox (reusing _detect_and_maybe_select_type) —
    this lets one monthly-block file per lab type be uploaded together and
    merged, instead of forcing everything into a single file.

    Duplicate Year+Type rows across files are summed.

    Returns (merged_df, col_map, errors, file_log) — same shape as
    generators.merge_files, so callers can treat single- and multi-file
    uploads identically.
    """
    frames = []
    file_log = []
    errors = []

    for i, path in enumerate(paths):
        fname = os.path.basename(path)
        file_types = _detect_and_maybe_select_type(path, allowed_types, f"{tool_key}_{i}")
        try:
            df_loaded, col_map, errs_l, warns_l = core.load_and_filter(path, file_types)
        except Exception as e:
            file_log.append(f"ERROR processing '{fname}': {e}")
            continue

        if errs_l:
            file_log.append(f"SKIPPED ({'; '.join(errs_l)}): {fname}")
            continue
        if df_loaded is None or df_loaded.empty:
            file_log.append(f"SKIPPED (empty/no matching types): {fname}")
            continue

        yc = col_map.get("year", "Year")
        tc = col_map.get("type", "Type")
        vc = col_map.get("value", "Value")
        df = df_loaded[[yc, tc, vc]].copy()
        df.columns = ["Year", "Type", "Value"]
        df["Value"] = pd.to_numeric(df["Value"], errors="coerce").fillna(0)
        df["Year"]  = pd.to_numeric(df["Year"], errors="coerce")
        df = df.dropna(subset=["Year", "Type", "Value"])
        df["Year"]  = df["Year"].astype(int)
        df["Type"]  = df["Type"].astype(str).str.strip()

        known = {t.lower(): t for t in allowed_types}
        df["Type"] = df["Type"].apply(lambda x: known.get(x.lower(), x))

        rows_kept = len(df)
        if rows_kept == 0:
            file_log.append(f"SKIPPED (no usable rows): {fname}")
            continue

        frames.append(df)
        log_line = f"OK ({rows_kept} rows): {fname}"
        if warns_l:
            log_line += f"  [{'; '.join(warns_l)}]"
        file_log.append(log_line)

    if not frames:
        errors.append("No valid data found in any uploaded file.")
        return None, {}, errors, file_log

    merged = pd.concat(frames, ignore_index=True)
    merged = merged.groupby(["Year", "Type"], as_index=False)["Value"].sum()

    found   = set(merged["Type"].unique())
    missing = set(allowed_types) - found
    if missing:
        file_log.append(f"WARNING: Lab types not found in any file: {sorted(missing)}")

    col_map_out = {"year": "Year", "type": "Type", "value": "Value"}
    return merged, col_map_out, errors, file_log

def get_annual_totals(weekly_df, types, years):
    """Return dict {year: {type: annual_total}}."""
    result = {}
    for y in years:
        yd = weekly_df[weekly_df["Year"] == y]
        result[int(y)] = {t: round(yd[t].sum(), 1) for t in types}
    return result

def util_badge(util):
    if util > 1.0:   return f'<span class="badge-red">🔴 Overloaded {util:.0%}</span>'
    if util >= 0.8:  return f'<span class="badge-yellow">🟡 Near Cap {util:.0%}</span>'
    return               f'<span class="badge-green">🟢 OK {util:.0%}</span>'

def generate_excel(tool_fn, *args, **kwargs):
    """Run generator in temp dir, return bytes of Excel file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        dummy_input = os.path.join(tmpdir, "input.xlsx")
        # Create a dummy input path so save_workbook knows where to save
        open(dummy_input, 'wb').close()
        out_path, warns = tool_fn(*args, **kwargs)
        with open(out_path, "rb") as f:
            return f.read(), warns


# ──────────────────────────────────────────────────────────
#  PLOTLY CHART BUILDERS
# ──────────────────────────────────────────────────────────

def chart_yoy_bar_and_pie(annual_totals, types, years, colors):
    """YoY grouped bar + single combined pie figure with per-pie color legends."""
    bar_colors = ["#1F3864","#2E75B6","#70AD47","#FFD700","#FF4444","#9DC3E6","#C6EFCE"]
    pie_colors = ["#2E75B6","#1F3864","#70AD47","#FFD700","#FF4444","#9DC3E6","#C6EFCE"]

    # ── Bar chart ─────────────────────────────────────────
    fig_bar = go.Figure()
    for i, t in enumerate(types):
        y_vals = [annual_totals[int(y)].get(t, 0) for y in years]
        fig_bar.add_trace(go.Bar(
            name=t, x=[str(y) for y in years], y=y_vals,
            marker_color=bar_colors[i % len(bar_colors)],
            text=[f"{v:.0f}" for v in y_vals], textposition="outside",
        ))
    n_types = len(types)
    # Divide by max(n_types,2) so single-type charts (Tool 3 / Tool 4 Thermal)
    # get the same bar width as Tool 1 (2 types) instead of an oversized 0.5
    bar_w = max(0.15, min(0.5, 0.6 / max(n_types, 2)))
    fig_bar.update_layout(
        barmode="group", title="Year-on-Year Demand by Lab Type",
        xaxis=dict(type="category", title="Year",
                   tickmode="array", tickvals=[str(y) for y in years],
                   ticktext=[str(y) for y in years]),
        yaxis_title="Samples/Year",
        legend=dict(orientation="h", yanchor="top", y=-0.15,
                    x=0.5, xanchor="center"),
        height=460, plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Arial", size=11), margin=dict(t=50, b=110, r=20),
        uniformtext=dict(minsize=8, mode="hide"), bargroupgap=0.12,
    )
    fig_bar.update_traces(width=bar_w)
    fig_bar.update_xaxes(showgrid=False)
    fig_bar.update_yaxes(showgrid=True, gridcolor="#F0F0F0")

    # ── Single combined pie figure ─────────────────────────
    n         = len(years)
    NCOLS     = min(n, 4)
    NROWS     = math.ceil(n / NCOLS)
    VS        = 0.28      # vertical spacing between rows — leaves room for legends
    HS        = 0.04

    specs = []
    for r in range(NROWS):
        row = []
        for c in range(NCOLS):
            row.append({"type": "pie"} if r * NCOLS + c < n else None)
        specs.append(row)

    fig_pies = make_subplots(
        rows=NROWS, cols=NCOLS,
        specs=specs,
        subplot_titles=[str(int(y)) for y in years],
        vertical_spacing=VS,
        horizontal_spacing=HS,
    )

    for i, year in enumerate(years):
        r = i // NCOLS + 1
        c = i %  NCOLS + 1
        vals = [annual_totals[int(year)].get(t, 0) for t in types]
        fig_pies.add_trace(
            go.Pie(
                labels=list(types), values=vals, name=str(int(year)),
                marker=dict(colors=pie_colors[:len(types)],
                            line=dict(color="#FFF", width=2)),
                textinfo="label+percent",
                textfont=dict(size=10),
                hovertemplate="<b>%{label}</b><br>%{value:.0f} samples<br>%{percent}<extra></extra>",
                showlegend=False,
            ),
            row=r, col=c,
        )

    # ── Per-pie legend annotations (colored ■ + text) ─────
    # Row heights in paper coords (row 1 = top)
    row_h = (1.0 - (NROWS - 1) * VS) / NROWS
    # Width each column occupies
    col_w = (1.0 - (NCOLS - 1) * HS) / NCOLS

    # Variable-width legend items: swatch + text width (scales with label length)
    # so long labels like "Cold Spray" don't overlap short ones like "HVOF"
    SWATCH_W = 0.020   # space reserved for the ■ glyph
    CHAR_W   = 0.0078  # approx width per character at font-size 10
    GAP_W    = 0.018   # trailing gap between legend items

    def _item_width(label):
        return SWATCH_W + len(label) * CHAR_W + GAP_W

    item_widths = [_item_width(t) for t in types]
    TOTAL_LEG_W = sum(item_widths)

    for i, year in enumerate(years):
        row_i = i // NCOLS   # 0-indexed
        col_i = i %  NCOLS   # 0-indexed

        # X-centre of this cell in paper coords
        x_ctr = col_i * (col_w + HS) + col_w * 0.5

        # Y-bottom of this row in paper coords
        # Row 0 (top) bottom = 1 - row_h
        # Row 1 (next) bottom = 1 - row_h - VS - row_h = 1 - 2*row_h - VS
        row_bottom = 1.0 - (row_i + 1) * row_h - row_i * VS
        y_leg = row_bottom - 0.04   # just below the pie

        # Starting x for first legend item (centred group)
        x_start = x_ctr - TOTAL_LEG_W / 2

        x_cursor = x_start
        for j, (t, col) in enumerate(zip(types, pie_colors)):
            x0 = x_cursor
            # Colored ■ swatch
            fig_pies.add_annotation(
                x=x0, y=y_leg,
                text="■",
                font=dict(color=col, size=14, family="Arial"),
                xref="paper", yref="paper",
                showarrow=False,
                xanchor="left",
            )
            # Type label
            fig_pies.add_annotation(
                x=x0 + SWATCH_W, y=y_leg,
                text=t,
                font=dict(color="#333", size=10, family="Arial"),
                xref="paper", yref="paper",
                showarrow=False,
                xanchor="left",
            )
            x_cursor += item_widths[j]

    # Bottom margin must be large enough for row NROWS annotations (y may be negative)
    bottom_margin = max(60, int(len(types) * 18 + 40))

    fig_pies.update_layout(
        title="Process-wise Demand Share (%) by Year",
        height=max(320, 300 * NROWS + bottom_margin),
        showlegend=False,
        paper_bgcolor="white",
        font=dict(family="Arial", size=11),
        margin=dict(t=50, b=bottom_margin, l=10, r=10),
    )

    return fig_bar, fig_pies


def chart_capacity_bar(weekly_df, types, capacities, years):
    """Capacity vs Avg Weekly Demand — demand bars + capacity reference line/marker."""
    fig = go.Figure()
    cap_weekly = {t: capacities.get(t, 1) / 52 for t in types}

    bar_colors = ["#1F3864","#2E75B6","#70AD47","#FF4444","#FFD700","#9DC3E6","#C6EFCE"]

    # ── Demand bars per year ──────────────────────────────
    # Bar width must scale with the number of BARS PER GROUP. In this
    # chart the x-axis groups by Lab Type, and each *year* contributes one
    # bar per group — so the bar count per group is len(years), not
    # len(types). The old formula divided by n_types instead, which left
    # bars far too wide whenever several years were loaded (e.g. 6 years
    # of 0.3-wide bars overlapping almost edge-to-edge). Basing it on
    # n_years fixes the oversized-bar look across Tools 1/2/3.
    n_years = len(years)
    bar_w = max(0.05, min(0.42, 0.75 / max(n_years, 2)))

    demand_vals = []
    for yi, year in enumerate(years):
        yd   = weekly_df[weekly_df["Year"] == year]
        avgs = [round(float(yd[t].mean()), 3) if t in yd.columns else 0 for t in types]
        demand_vals.extend(v for v in avgs if not math.isnan(v))
        fig.add_trace(go.Bar(
            name=f"{year} Avg Demand",
            x=types, y=avgs,
            marker_color=bar_colors[yi % len(bar_colors)],
            text=[f"{v:.2f}" for v in avgs],
            textposition="outside",
            textfont=dict(size=9),
            width=bar_w,
        ))

    # ── Capacity: scatter markers (▬) at cap level per type ──
    cap_vals = [round(cap_weekly[t], 2) for t in types]
    fig.add_trace(go.Scatter(
        name="Weekly Capacity",
        x=types,
        y=cap_vals,
        mode="markers+text",
        marker=dict(
            symbol="line-ew",
            size=36,
            color="#C00000",
            line=dict(width=3, color="#C00000"),
        ),
        text=[f"Cap: {v:.2f}" for v in cap_vals],
        textposition="top center",
        textfont=dict(size=9, color="#C00000"),
    ))

    # Y range: headroom above the larger of capacity or peak demand
    safe_demand = [v for v in demand_vals if not math.isnan(v)]
    y_top = max(
        max(safe_demand, default=0),
        max(cap_vals, default=0)
    ) * 1.35

    fig.update_layout(
        barmode="group",
        title="Average Weekly Demand vs Capacity",
        xaxis=dict(type="category", title="Lab Type"),
        yaxis=dict(title="Weekly Units", range=[0, y_top]),
        height=460, plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Arial", size=11),
        legend=dict(orientation="h", yanchor="top", y=-0.15, x=0),
        margin=dict(b=100, t=60, r=20),
        uniformtext=dict(minsize=8, mode="hide"),
        bargap=0.3, bargroupgap=0.05,
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(showgrid=True, gridcolor="#F0F0F0")
    return fig


def chart_utilization_line(util_df, types, years, colors):
    """Monthly utilization line chart with 100%/80% reference lines."""
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    line_colors = ["#1F3864","#2E75B6","#70AD47","#FF4444","#FFD700","#9DC3E6"]

    fig = go.Figure()
    for yi, year in enumerate(years):
        yd = util_df[util_df["Year"] == year].copy()
        yd["_m"] = yd["Week"].apply(core.week_to_month)
        for ti, t in enumerate(types):
            monthly = [yd[yd["_m"] == m+1][t].mean() for m in range(12)]
            monthly = [v if not math.isnan(v) else 0 for v in monthly]
            fig.add_trace(go.Scatter(
                x=months, y=monthly,
                name=f"{t} ({year})",
                line=dict(color=line_colors[ti % len(line_colors)],
                          width=2,
                          dash="solid" if yi == 0 else "dash"),
                mode="lines+markers",
                marker=dict(size=6),
                hovertemplate=f"<b>{t} ({year})</b><br>%{{x}}: %{{y:.1%}}<extra></extra>",
            ))

    fig.add_hline(y=1.0, line_dash="dot", line_color="#FF4444",
                  annotation_text="100% Capacity", annotation_position="right")
    fig.add_hline(y=0.8, line_dash="dot", line_color="#FFD700",
                  annotation_text="80% Threshold", annotation_position="right")

    # ── Dynamic legend sizing — avoids overlap with axis title
    # when there are many year × type combinations ──────────
    n_entries     = len(years) * len(types)
    entries_per_row = 6
    legend_rows   = max(1, math.ceil(n_entries / entries_per_row))
    legend_y      = -0.18 - (legend_rows - 1) * 0.09
    margin_b      = 90 + legend_rows * 38
    chart_height  = 440 + max(0, legend_rows - 2) * 40

    fig.update_layout(
        title="Monthly Utilization Trend by Lab Type",
        xaxis_title=None,   # "Month" removed — labels (Jan–Dec) are self-explanatory
                            # and prevented overlap with the multi-row legend below
        yaxis_title="Utilization",
        yaxis_tickformat=".0%",
        height=chart_height, plot_bgcolor="white", paper_bgcolor="white",
        font=dict(family="Arial", size=12),
        legend=dict(orientation="h", yanchor="top", y=legend_y,
                    x=0.5, xanchor="center",
                    font=dict(size=10)),
        margin=dict(b=margin_b),
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(showgrid=True, gridcolor="#F0F0F0")
    return fig


def chart_gantt(weekly_df, types, capacities, year, current_week):
    """Gantt heatmap for a given year."""
    weeks = list(range(1, 53))
    yd    = weekly_df[weekly_df["Year"] == year]

    z, labels, text_grid = [], [], []
    for i, t in enumerate(types):
        cap_wk = capacities.get(t, 1) / 52
        row_z, row_txt = [], []
        for w in weeks:
            wrow = yd[yd["Week"] == w]
            dem  = float(wrow[t].values[0]) if not wrow.empty else 0.0
            util = dem / cap_wk if cap_wk > 0 else 0.0
            # clamp to [0, 1.5] for display but store raw for tooltip
            row_z.append(round(min(util, 1.5), 3))
            row_txt.append(f"Week {w} | {t}<br>Demand: {dem:.2f}<br>Util: {util:.1%}")
        z.append(row_z)
        labels.append(t)
        text_grid.append(row_txt)

    # Colorscale values MUST be 0.0–1.0 (plotly requirement)
    # Map: 0→green, 0.53→yellow (=80% of 1.5 scale), 0.67→red (=100%), 1.0→dark red
    colorscale = [
        [0.00, "#C6EFCE"],   # 0% util → light green
        [0.40, "#70AD47"],   # ~60% util → green
        [0.53, "#FFD700"],   # ~80% util → yellow
        [0.67, "#FF4444"],   # ~100% util → red
        [1.00, "#CC0000"],   # ≥150% util → dark red
    ]

    fig = go.Figure(go.Heatmap(
        z=z,
        x=[f"Wk{w}" for w in weeks],
        y=labels,
        colorscale=colorscale,
        zmin=0, zmax=1.5,
        text=text_grid,
        hovertemplate="%{text}<extra></extra>",
        colorbar=dict(
            title=dict(text="Utilization", side="right"),
            tickvals=[0, 0.4, 0.8, 1.0, 1.5],
            ticktext=["0%", "40%", "80%", "100%", "≥150%"],
            len=0.8,
        ),
        showscale=True,
    ))

    # Mark current week
    if year == CURRENT_YEAR and 1 <= current_week <= 52:
        fig.add_vline(
            x=current_week - 1,          # vline on x-axis uses index not label
            line_color="#C00000", line_width=2, line_dash="dash",
            annotation_text=f"Wk {current_week} (Now)",
            annotation_position="top right",
            annotation_font_size=10,
        )

    fig.update_layout(
        title=f"Gantt Heatmap — {year} Weekly Occupancy",
        xaxis_title="Week",
        yaxis_title="Lab Type",
        height=max(280, 90 * len(types) + 140),
        font=dict(family="Arial", size=11),
        paper_bgcolor="white",
        plot_bgcolor="white",
        margin=dict(l=110, r=80, t=60, b=60),
    )
    return fig


def chart_comparison_grouped(annual_a, annual_b, types_a, types_b, years,
                              cap_a_total, cap_b_total):
    """Grouped bar: both groups side by side per year."""
    fig = make_subplots(rows=1, cols=2,
                        subplot_titles=[GROUP_A["name"], GROUP_B["name"]])

    col_map_a = ["#1F3864","#2E75B6","#9DC3E6"]
    col_map_b = ["#1F5C1A","#70AD47","#C6EFCE"]

    for ti, t in enumerate(types_a):
        vals = [annual_a[int(y)].get(t, 0) for y in years]
        fig.add_trace(go.Bar(name=t, x=[str(y) for y in years], y=vals,
                             marker_color=col_map_a[ti % 3],
                             legendgroup="A",
                             text=[f"{v:.0f}" for v in vals],
                             textposition="outside"), row=1, col=1)

    for ti, t in enumerate(types_b):
        vals = [annual_b[int(y)].get(t, 0) for y in years]
        fig.add_trace(go.Bar(name=t, x=[str(y) for y in years], y=vals,
                             marker_color=col_map_b[ti % 3],
                             legendgroup="B",
                             text=[f"{v:.0f}" for v in vals],
                             textposition="outside"), row=1, col=2)

    # Capacity reference lines + corner badges (see chart_comparison_3groups
    # for why these are pinned to fixed paper-domain corners rather than
    # attached to the line itself — avoids overlap with bar value labels).
    def _cap_badge(row, col, cap_val, color, label):
        fig.add_hline(y=cap_val, line_dash="dot", line_color=color, row=row, col=col)
        fig.add_annotation(
            text=f"{label}: {cap_val:.0f}/yr",
            xref="x domain", yref="y domain",
            x=0.02, y=0.97, xanchor="left", yanchor="top",
            showarrow=False,
            font=dict(size=10, color=color),
            bgcolor="rgba(255,255,255,0.88)",
            bordercolor=color, borderwidth=1, borderpad=3,
            row=row, col=col,
        )

    _cap_badge(1, 1, cap_a_total, "#1A4E8A", "Cap A")
    _cap_badge(1, 2, cap_b_total, "#1F5C1A", "Cap B")

    fig.update_layout(barmode="group", height=440,
                      font=dict(family="Arial", size=12),
                      paper_bgcolor="white", plot_bgcolor="white",
                      title="Annual Demand Comparison — Mechanical vs Coating Labs",
                      margin=dict(b=60))
    return fig


def chart_util_comparison_line(wdf_a, wdf_b, types_a, types_b, years,
                                cap_a_total, cap_b_total):
    """Line chart: group-level utilization over time."""
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    fig = go.Figure()

    for yi, year in enumerate(list(years)[-3:]):  # last 3 years for clarity
        # Group A
        yd_a = wdf_a[wdf_a["Year"] == year].copy()
        yd_a["_m"] = yd_a["Week"].apply(core.week_to_month)
        util_a = []
        for m in range(12):
            tot = sum(yd_a[yd_a["_m"]==m+1][t].mean()
                      for t in types_a if t in yd_a.columns)
            weekly_cap = cap_a_total / 52
            util_a.append(tot / weekly_cap if weekly_cap > 0 and not math.isnan(tot) else 0)

        fig.add_trace(go.Scatter(
            x=months, y=util_a, name=f"Mechanical {year}",
            line=dict(color="#1A4E8A", width=2,
                      dash="solid" if yi == 0 else "dash"),
            mode="lines+markers", marker=dict(size=5),
        ))

        # Group B
        yd_b = wdf_b[wdf_b["Year"] == year].copy()
        yd_b["_m"] = yd_b["Week"].apply(core.week_to_month)
        util_b = []
        for m in range(12):
            tot = sum(yd_b[yd_b["_m"]==m+1][t].mean()
                      for t in types_b if t in yd_b.columns)
            weekly_cap = cap_b_total / 52
            util_b.append(tot / weekly_cap if weekly_cap > 0 and not math.isnan(tot) else 0)

        fig.add_trace(go.Scatter(
            x=months, y=util_b, name=f"Coating {year}",
            line=dict(color="#1F5C1A", width=2,
                      dash="solid" if yi == 0 else "dash"),
            mode="lines+markers", marker=dict(size=5),
        ))

    fig.add_hline(y=1.0, line_dash="dot", line_color="#FF4444",
                  annotation_text="100%")
    fig.add_hline(y=0.8, line_dash="dot", line_color="#FFD700",
                  annotation_text="80%")

    fig.update_layout(
        title="Monthly Utilization % — Mechanical vs Coating Labs (last 3 years)",
        xaxis_title="Month", yaxis_title="Utilization",
        yaxis_tickformat=".0%", height=420,
        font=dict(family="Arial", size=12),
        paper_bgcolor="white", plot_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=-0.4),
        margin=dict(b=120),
    )
    return fig


# ──────────────────────────────────────────────────────────
#  SHARED UI COMPONENTS
# ──────────────────────────────────────────────────────────

def show_kpi_row(types, annual_totals, capacities, years):
    """Show KPI metric cards for most recent year, including vacant capacity."""
    last_year = max(years)
    cols = st.columns(len(types))
    for ci, t in enumerate(types):
        cap    = capacities.get(t, 1)
        dem    = annual_totals[int(last_year)].get(t, 0)
        util   = dem / cap if cap > 0 else 0
        vacant = max(0, cap - dem)
        with cols[ci]:
            cls  = "red" if util > 1 else "amber" if util >= 0.8 else "green"
            icon = "🔴" if util > 1 else "🟡" if util >= 0.8 else "🟢"
            vacant_color = "#888" if vacant > 0 else "#FF4444"
            st.markdown(f"""
            <div class="metric-card {cls}">
                <div style="font-size:11px;color:#666;font-weight:600;">{t}</div>
                <div style="font-size:22px;font-weight:800;color:#1F3864;">{dem:.0f}</div>
                <div style="font-size:12px;color:#555;">samples in {last_year}</div>
                <div style="font-size:13px;margin-top:4px;">{icon} {util:.0%} of {cap}/yr</div>
                <div style="font-size:12px;color:{vacant_color};margin-top:2px;">
                    🕳️ Vacant: <b>{vacant:.0f}</b> slots/yr ({vacant/52:.1f}/wk)
                </div>
            </div>
            """, unsafe_allow_html=True)


def show_util_table(util_df, types, years):
    """Colored utilization summary table."""
    rows = []
    for t in types:
        row = {"Lab Type": t}
        for y in years:
            peak = util_df[util_df["Year"] == y][t].max()
            avg  = util_df[util_df["Year"] == y][t].mean()
            row[f"{y} Peak"] = f"{peak:.1%}"
            row[f"{y} Avg"]  = f"{avg:.1%}"
        rows.append(row)
    df_display = pd.DataFrame(rows)

    def color_util(val):
        if isinstance(val, str) and val.endswith("%"):
            try:
                v = float(val.replace("%","").replace("+","")) / 100
                if v > 1.0: return "background-color:#FFCCCC; color:#CC0000; font-weight:bold"
                if v >= 0.8: return "background-color:#FFF3CC"
                return "background-color:#CCEFCC"
            except: pass
        return ""

    st.dataframe(
        df_display.style.map(color_util, subset=[c for c in df_display.columns if c != "Lab Type"]),
        use_container_width=True, hide_index=True,
    )
    st.markdown("""
    <div style="display:flex;gap:20px;align-items:center;flex-wrap:wrap;
                margin:4px 0 14px 2px;font-size:12px;color:#444;">
        <span style="font-weight:600;color:#666;">Legend:</span>
        <span><span style="display:inline-block;width:12px;height:12px;
              background:#CCEFCC;border:1px solid #8FBF8F;border-radius:2px;
              margin-right:5px;vertical-align:middle;"></span>Under 80% — within capacity</span>
        <span><span style="display:inline-block;width:12px;height:12px;
              background:#FFF3CC;border:1px solid #D8C27A;border-radius:2px;
              margin-right:5px;vertical-align:middle;"></span>80–100% — near capacity</span>
        <span><span style="display:inline-block;width:12px;height:12px;
              background:#FFCCCC;border:1px solid #E08A8A;border-radius:2px;
              margin-right:5px;vertical-align:middle;"></span>Over 100% — over capacity</span>
    </div>
    """, unsafe_allow_html=True)


def show_annual_summary_table(annual_totals, types, years, capacities):
    """
    Annual summary table — mirrors Image 3.
    Columns: Year | Lab Type | Annual Demand | Capacity | Utilization (%) | Status | Vacancy/yr | Vacancy/wk
    Colour-coded: red >100%, amber 80-100%, green <80%
    """
    rows = []
    for yr in sorted([int(y) for y in years]):
        for t in types:
            dem    = annual_totals.get(yr, {}).get(t, 0)
            cap    = capacities.get(t, 1)
            util   = dem / cap * 100 if cap > 0 else 0
            vacant = max(0.0, cap - dem)
            if util > 100:
                status = "OVER CAPACITY"
            elif util >= 80:
                status = "HIGH"
            else:
                status = "OK"
            rows.append({
                "Year":               yr,
                "Lab Type":           t,
                "Annual Demand":      int(round(dem)),
                "Capacity":           int(cap),
                "Utilization (%)":    f"{util:.1f}%",
                "Status":             status,
                "Vacancy (slots/yr)": int(vacant),
                "Vacancy (slots/wk)": f"{vacant / 52:.1f}",
            })

    df = pd.DataFrame(rows)

    def _style_util(v):
        if not isinstance(v, str) or "%" not in v:
            return ""
        pct = float(v.replace("%", ""))
        if pct > 100:   return "background-color:#FF4444;color:white;font-weight:700;text-align:center"
        if pct >= 80:   return "background-color:#FFA500;color:white;font-weight:700;text-align:center"
        return                  "background-color:#70AD47;color:white;font-weight:700;text-align:center"

    def _style_status(v):
        if v == "OVER CAPACITY": return "background-color:#FF4444;color:white;font-weight:700"
        if v == "HIGH":           return "background-color:#FFA500;color:white;font-weight:700"
        return                           "background-color:#70AD47;color:white;font-weight:700"

    def _style_vacancy(v):
        if isinstance(v, (int, float)) and v == 0:
            return "background-color:#FFD7D7;color:#CC0000;font-weight:700"
        return "background-color:#E2EFDA;color:#375623"

    st.markdown("##### 📋 Annual Summary")
    styled = (
        df.style
          .map(_style_util,    subset=["Utilization (%)"])
          .map(_style_status,  subset=["Status"])
          .map(_style_vacancy, subset=["Vacancy (slots/yr)"])
    )
    st.dataframe(styled, use_container_width=True, hide_index=True)


def make_editable_excel(weekly_df, util_df, annual_totals, types, capacities, years, title="Lab Data"):
    """Build an editable Excel workbook with 3 sheets and return bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def hdr_cell(ws, row, col, val, bg="1F3864", fg="FFFFFF", sz=10, bold=True):
        c = ws.cell(row, col, val)
        c.font = Font(name="Arial", size=sz, bold=bold, color=fg)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin
        return c

    def data_cell(ws, row, col, val, fmt=None, bg=None):
        c = ws.cell(row, col, val)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = thin
        if fmt: c.number_format = fmt
        if bg:  c.fill = PatternFill("solid", start_color=bg)
        return c

    # ── Sheet 1: Weekly Demand (editable) ──────────────────
    ws1 = wb.active
    ws1.title = "Weekly_Demand"
    ws1.cell(1, 1, f"📋 {title} — Editable Weekly Demand").font = Font(
        name="Arial", size=12, bold=True, color="1F3864")
    ws1.cell(2, 1, "✏️  You can edit the values in this sheet. Yellow = editable cells.")
    ws1.cell(2, 1).font = Font(name="Arial", size=9, italic=True, color="7F6000")

    # Header row
    headers = ["Year", "Week"] + list(types)
    for ci, h in enumerate(headers, 1):
        bg = "1F3864" if ci <= 2 else "2E75B6"
        hdr_cell(ws1, 4, ci, h, bg=bg)
        ws1.column_dimensions[get_column_letter(ci)].width = 14 if ci > 2 else 8

    editable_fill = PatternFill("solid", start_color="FFFFE0")  # light yellow
    for ri, row in enumerate(weekly_df.itertuples(index=False), 5):
        vals = list(row)
        data_cell(ws1, ri, 1, vals[0])  # Year
        data_cell(ws1, ri, 2, vals[1])  # Week
        for ci, t in enumerate(types, 3):
            c = data_cell(ws1, ri, ci, round(vals[2 + list(weekly_df.columns[2:]).index(t)], 3),
                          fmt="0.000")
            c.fill = editable_fill  # mark as editable

    ws1.freeze_panes = "A5"
    ws1.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    # ── Sheet 2: Annual Summary ─────────────────────────────
    ws2 = wb.create_sheet("Annual_Summary")
    ws2.cell(1, 1, f"📊 {title} — Annual Summary").font = Font(
        name="Arial", size=12, bold=True, color="1F3864")

    hdrs2 = ["Lab Type", "Capacity/yr", "Weekly Cap"] + [str(y) for y in years] + ["Best Year", "Worst Year"]
    for ci, h in enumerate(hdrs2, 1):
        hdr_cell(ws2, 3, ci, h)
        ws2.column_dimensions[get_column_letter(ci)].width = 14

    for ri, t in enumerate(types, 4):
        cap = capacities.get(t, 1)
        data_cell(ws2, ri, 1, t)
        data_cell(ws2, ri, 2, cap, fmt="#,##0")
        data_cell(ws2, ri, 3, round(cap / 52, 2), fmt="0.00")
        yr_vals = {}
        for ci, y in enumerate(years, 4):
            dem = annual_totals.get(int(y), {}).get(t, 0)
            yr_vals[y] = dem
            util = dem / cap if cap > 0 else 0
            bg = "FF4444" if util > 1 else "FFD700" if util >= 0.8 else "C6EFCE"
            c = data_cell(ws2, ri, ci, round(dem, 1), fmt="#,##0.0", bg=bg)
            if util > 1:
                c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        best_yr = max(yr_vals, key=yr_vals.get, default="")
        worst_yr = min(yr_vals, key=yr_vals.get, default="")
        data_cell(ws2, ri, len(hdrs2), str(worst_yr))
        data_cell(ws2, ri, len(hdrs2) - 1, str(best_yr))

    ws2.freeze_panes = "A4"

    # ── Sheet 3: Utilization Summary ───────────────────────
    ws3 = wb.create_sheet("Utilization_%")
    ws3.cell(1, 1, f"📈 {title} — Utilization % (Green<80% | Yellow 80-100% | Red>100%)").font = Font(
        name="Arial", size=11, bold=True, color="1F3864")

    hdrs3 = ["Lab Type"] + [f"{y} Peak" for y in years] + [f"{y} Avg" for y in years]
    for ci, h in enumerate(hdrs3, 1):
        hdr_cell(ws3, 3, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = 13

    for ri, t in enumerate(types, 4):
        data_cell(ws3, ri, 1, t)
        cap = capacities.get(t, 1)
        for ci, y in enumerate(years, 2):
            peak = util_df[util_df["Year"] == y][t].max() if t in util_df.columns else 0
            avg  = util_df[util_df["Year"] == y][t].mean() if t in util_df.columns else 0
            for val, offset in [(peak, 0), (avg, len(years))]:
                col = ci + offset
                bg = "FF4444" if val > 1 else "FFD700" if val >= 0.8 else "C6EFCE"
                c = data_cell(ws3, ri, col, round(val, 4), fmt="0.0%", bg=bg)
                if val > 1:
                    c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    ws3.freeze_panes = "B4"

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def capacity_warning(individual_caps, combined_cap):
    """Show warning if individual caps sum > combined cap."""
    total = sum(individual_caps.values())
    if total > combined_cap:
        st.warning(
            f"⚠️ Individual capacities sum to **{total}**, which exceeds "
            f"the combined cap of **{combined_cap}**. "
            "Consider reducing individual values or raising the combined cap."
        )
    elif total < combined_cap:
        st.info(
            f"ℹ️ Individual capacities sum to **{total}** "
            f"(combined cap is **{combined_cap}**). "
            f"**{combined_cap - total}** capacity units are unallocated."
        )
    else:
        st.success(f"✅ Individual capacities sum exactly to combined cap ({combined_cap}).")


# ──────────────────────────────────────────────────────────
#  SESSION STATE — Data Chat
# ──────────────────────────────────────────────────────────
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
for _tk in ("data_tool1", "data_tool2", "data_tool3"):
    if _tk not in st.session_state:
        st.session_state[_tk] = None
# Per-tool file caches — survive widget-triggered re-runs
for _fk in ("_fc_t1", "_fc_t2", "_fc_t3", "_fc_t4"):
    if _fk not in st.session_state:
        st.session_state[_fk] = None


# ──────────────────────────────────────────────────────────
#  SIDEBAR
# ──────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 🔬 Lab Dashboard Suite")
    st.markdown("---")
    tool = st.radio(
        "Select Tool",
        ["🔵 Tool 1 — LCF & Creep",
         "🟢 Tool 2 — Coating Labs",
         "🟠 Tool 3 — Thermal Lab",
         "🔴 Tool 4 — Comparison & PPT"],
        label_visibility="collapsed",
    )
    st.caption(f"Week **{CURRENT_WEEK}** / {CURRENT_YEAR}")
    with st.expander("ℹ️ Input format & output"):
        st.markdown("**Input:** Excel with `Year | Type | Value`")
        st.markdown("**Output:** Excel dashboard (7–8 sheets) + optional PowerPoint")
    st.markdown("---")

    # ── Rule-Based Data Assistant ──────────────────────────
    st.markdown("### 💬 Data Assistant")
    st.caption("Ask questions about your uploaded lab data.")

    # ── answer engine ──────────────────────────────────────
    def _find_lab_type(question, all_types):
        q = question.lower()
        for t in sorted(all_types, key=len, reverse=True):
            if all(w in q for w in t.lower().split()):
                return t
        return None

    def _find_year_in_q(question):
        import re as _re
        m = _re.search(r'\b(20\d{2})\b', question)
        return int(m.group(1)) if m else None

    def answer_data_question(question: str, session_data: dict) -> str:
        import math as _math
        GROUP_LABEL_MAP = {
            "data_tool1": "Mechanical (Tool 1)",
            "data_tool2": "Coating (Tool 2)",
            "data_tool3": "Thermal (Tool 3)",
        }
        if not session_data:
            return ("No data loaded yet. Upload a file in any tool tab first, "
                    "then come back and ask a question.")
        q = question.lower().strip()
        all_types_map = {}
        for tk, td in session_data.items():
            for t in td.get("types", []):
                all_types_map[t] = tk
        all_types = list(all_types_map.keys())

        # 2 — utilization
        if any(kw in q for kw in ["utilization", "utilisation", "util"]):
            lab  = _find_lab_type(q, all_types)
            year = _find_year_in_q(q)
            if lab:
                tk      = all_types_map[lab]
                udf     = session_data[tk]["util_df"]
                yr_data = udf[udf["Year"] == year] if year else udf
                if yr_data.empty or lab not in yr_data.columns:
                    return f"No utilization data found for **{lab}**{' in '+str(year) if year else ''}."
                peak   = yr_data[lab].max()
                avg    = yr_data[lab].mean()
                status = "🔴 OVERLOADED" if peak > 1 else "🟡 Near Cap" if peak >= 0.8 else "🟢 Healthy"
                yr_str = f" in {year}" if year else " (all years)"
                return (f"**{lab}** utilization{yr_str}:\n"
                        f"• Peak: **{peak:.1%}** | Avg: **{avg:.1%}**\n"
                        f"• Status: {status}")
            lines = ["Utilization across all loaded labs:"]
            for t in all_types:
                tk  = all_types_map[t]; udf = session_data[tk]["util_df"]
                if t not in udf.columns: continue
                p   = udf[t].max(); a = udf[t].mean()
                st2 = "🔴" if p > 1 else "🟡" if p >= 0.8 else "🟢"
                lines.append(f"  {st2} **{t}**: peak {p:.1%} | avg {a:.1%}")
            return "\n".join(lines)

        # 3 — highest demand
        if any(kw in q for kw in ["highest demand","most demand","peak demand",
                                   "max demand","most samples"]):
            lab = _find_lab_type(q, all_types)
            results = []
            for t in ([lab] if lab else all_types):
                tk = all_types_map[t]; annual = session_data[tk]["annual"]
                best_yr, best_val = None, -1
                for yr, vals in annual.items():
                    v = vals.get(t, 0)
                    if v > best_val: best_val, best_yr = v, yr
                if best_yr is not None: results.append((t, best_yr, best_val))
            if not results: return "No demand data found."
            results.sort(key=lambda x: x[2], reverse=True)
            lines = ["**Highest demand by lab:**"]
            for t, yr, val in results[:6]:
                lines.append(f"  • **{t}**: {yr} — **{val:,.0f}** samples")
            return "\n".join(lines)

        # 4 — most overloaded
        if any(kw in q for kw in ["overload","over capacity","highest utilization",
                                   "highest utilisation","most stressed",
                                   "most utilized","most utilised"]):
            best = None
            for t in all_types:
                tk = all_types_map[t]; udf = session_data[tk]["util_df"]
                if t not in udf.columns: continue
                for yr in udf["Year"].unique():
                    p = udf[udf["Year"]==yr][t].max()
                    if best is None or p > best[2]: best = (t, int(yr), p)
            if best is None: return "No utilization data found."
            t, yr, p = best
            status = "🔴 OVERLOADED" if p > 1 else "🟡 Near Cap" if p >= 0.8 else "🟢 Healthy"
            return (f"Most loaded lab: **{t}** in **{yr}**\n"
                    f"• Peak utilization: **{p:.1%}** — {status}")

        # 5 — capacity for 80%
        if any(kw in q for kw in ["capacity needed","capacity for 80",
                                   "need for 80","80% capacity","80 percent",
                                   "capacity increase","how much capacity"]):
            lab = _find_lab_type(q, all_types)
            if not lab:
                return ("Please name a lab type. Example:\n"
                        "'What capacity do I need for 80% in Cold Spray?'")
            tk     = all_types_map[lab]
            annual = session_data[tk]["annual"]
            cap    = session_data[tk]["capacities"].get(lab, 0)
            yrs    = sorted(annual.keys(), reverse=True)
            if not yrs: return f"No demand data found for **{lab}**."
            last_yr = yrs[0]
            demand  = annual[last_yr].get(lab, 0)
            needed  = _math.ceil(demand / 0.8)
            util_now = demand / cap if cap > 0 else 0
            delta    = needed - cap
            arrow    = f"↑ +{delta}" if delta > 0 else "✅ already sufficient"
            return (f"**{lab}** capacity planning ({last_yr}):\n"
                    f"• Demand: **{demand:,.0f}** samples\n"
                    f"• Current capacity: **{cap:,}**/yr ({util_now:.1%} utilization)\n"
                    f"• Needed for ≤80% util: **{needed:,}**/yr ({arrow})")

        # 6 — list labs
        if any(kw in q for kw in ["which lab","list lab","what lab",
                                   "show lab","all lab","available lab"]):
            lines = ["**Labs currently loaded:**"]
            for tk, td in session_data.items():
                label = GROUP_LABEL_MAP.get(tk, tk)
                types = td.get("types", []); yrs = td.get("years", [])
                caps  = td.get("capacities", {})
                yr_str = f"{min(yrs)}–{max(yrs)}" if yrs else "—"
                lines.append(f"\n**{label}** ({yr_str})")
                for t in types:
                    lines.append(f"  • {t}  ({caps.get(t,'?')}/yr)")
            return "\n".join(lines)

        # 7 — summary
        if any(kw in q for kw in ["summary","overview","give me a summary",
                                   "show summary","what do i have"]):
            lines = ["**Dashboard summary:**"]
            for tk, td in session_data.items():
                label  = GROUP_LABEL_MAP.get(tk, tk)
                types  = td.get("types", []); yrs = td.get("years", [])
                annual = td.get("annual", {}); caps = td.get("capacities", {})
                udf    = td.get("util_df", None)
                if not yrs: continue
                last_yr = max(yrs)
                dem  = sum(annual.get(last_yr, {}).get(t, 0) for t in types)
                cap  = sum(caps.values())
                util = dem / cap if cap > 0 else 0
                peak = 0
                if udf is not None:
                    yr_udf = udf[udf["Year"] == last_yr]
                    for t in types:
                        if t in yr_udf.columns:
                            v = yr_udf[t].max()
                            if v > peak: peak = v
                status = "🔴" if peak > 1 else "🟡" if peak >= 0.8 else "🟢"
                lines.append(
                    f"\n{status} **{label}** — {', '.join(types)}\n"
                    f"   {last_yr}: demand **{dem:,.0f}** | util **{util:.0%}** | peak **{peak:.0%}**"
                )
            return "\n".join(lines)

        # 8 — trend
        if any(kw in q for kw in ["trend","annual trend","year by year",
                                   "over the years","each year","by year"]):
            lab = _find_lab_type(q, all_types)
            if not lab:
                lines = ["**Annual demand trends:**"]
                for t in all_types:
                    tk     = all_types_map[t]; annual = session_data[tk]["annual"]
                    vals   = " | ".join(f"{y}: {annual.get(y,{}).get(t,0):.0f}"
                                        for y in sorted(annual.keys()))
                    lines.append(f"  • **{t}**: {vals}")
                return "\n".join(lines)
            tk     = all_types_map[lab]; annual = session_data[tk]["annual"]
            vals   = " | ".join(f"{y}: {annual.get(y,{}).get(lab,0):.0f}"
                                for y in sorted(annual.keys()))
            return f"**{lab}** annual demand:\n  {vals}"

        # 9 — unrecognized
        return (
            "I can only answer questions about uploaded data. Try:\n"
            "  • *'LCF utilization in 2024'*\n"
            "  • *'Which lab is most overloaded?'*\n"
            "  • *'Capacity needed for 80% in Cold Spray'*\n"
            "  • *'Show me a summary'*\n"
            "  • *'Highest demand for Plasma'*\n"
            "  • *'Trend for Thermal Rig'*\n"
            "  • *'List all labs'*"
        )

    # ── build session_data dict from loaded tools ──────────
    def _build_session_data():
        sd = {}
        for tk in ("data_tool1", "data_tool2", "data_tool3"):
            val = st.session_state.get(tk)
            if val is not None:
                sd[tk] = val
        return sd

    # ── chat bubble display ────────────────────────────────
    bubbles_html = '<div class="chat-wrap">'
    for msg in st.session_state.chat_history[-14:]:
        if msg["role"] == "user":
            bubbles_html += f'<div class="chat-user">🧑 {msg["content"]}</div>'
        else:
            txt = msg["content"].replace("\n", "<br>").replace("**", "<b>", 1)
            # Bold markdown: simplistic pass
            import re as _re2
            txt = _re2.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>',
                           msg["content"].replace("\n", "<br>"))
            bubbles_html += f'<div class="chat-bot">💬 {txt}</div>'
    bubbles_html += "</div>"
    st.markdown(bubbles_html, unsafe_allow_html=True)

    # ── input + buttons ────────────────────────────────────
    user_question = st.text_input(
        "Ask:", key="chat_input", label_visibility="collapsed",
        placeholder="e.g. LCF utilization in 2024 | capacity for 80% Plasma | summary",
    )
    col_ask, col_clr = st.columns(2)
    with col_ask:
        ask_clicked = st.button("Ask", key="chat_ask", use_container_width=True)
    with col_clr:
        if st.button("🗑️ Clear", key="chat_clear", use_container_width=True):
            st.session_state.chat_history = []
            st.rerun()

    if ask_clicked and user_question.strip():
        q_text = user_question.strip()
        sd     = _build_session_data()
        answer = answer_data_question(q_text, sd)
        st.session_state.chat_history.append({"role": "user",    "content": q_text})
        st.session_state.chat_history.append({"role": "assistant","content": answer})
        st.rerun()


# ══════════════════════════════════════════════════════════════════
#  TOOL 1 — LCF & CREEP
# ══════════════════════════════════════════════════════════════════

if "Tool 1" in tool:
    st.markdown("""
    <div class="tool-header">
        <h2>🔵 LCF & Creep Lab Dashboard</h2>
        <p>Mechanical lab occupancy planning — individual capacity tracking</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Config ────────────────────────────────────────────
    col_a, col_b, col_c = st.columns([2, 1, 1])
    with col_a:
        st.markdown('<div class="section-label">📂 Upload Excel File(s)</div>', unsafe_allow_html=True)
        uploaded_1 = st.file_uploader(
            "", type=["xlsx", "xls"], key="t1_file",
            accept_multiple_files=True, label_visibility="collapsed",
            help="Upload one combined file, or one file per lab type (e.g. one "
                 "monthly-block file for LCF and another for Creep) — they'll be merged.",
        )
        fc1_list = _persist_upload_multi("_fc_t1", uploaded_1 or [])
        if fc1_list and not uploaded_1:
            st.success(f"✅ {len(fc1_list)} file(s) loaded: {', '.join(f['name'] for f in fc1_list)}")
        elif uploaded_1:
            st.success(f"✅ {len(uploaded_1)} file(s) loaded: {', '.join(f.name for f in uploaded_1)}")
        if fc1_list:
            if st.button("✖ Remove file(s)", key="t1_remove"):
                st.session_state["_fc_t1"] = None
                st.rerun()
    with col_b:
        st.markdown('<div class="section-label">LCF Capacity</div>', unsafe_allow_html=True)
        lcf_cap = st.number_input("LCF (samples/yr)", value=50, min_value=1, max_value=9999,
                                   key="t1_lcf", label_visibility="collapsed")
        st.caption("LCF — samples/year")
    with col_c:
        st.markdown('<div class="section-label">Creep Capacity</div>', unsafe_allow_html=True)
        creep_cap = st.number_input("Creep (samples/yr)", value=22, min_value=1, max_value=9999,
                                     key="t1_creep", label_visibility="collapsed")
        st.caption("Creep — samples/year")

    theme_name = st.selectbox("Color Theme", list(core.COLOR_THEMES.keys()), key="t1_theme")
    caps_1 = {"LCF": lcf_cap, "Creep": creep_cap}
    theme_1 = core.COLOR_THEMES[theme_name]

    if fc1_list:
        paths_1 = [_bytes_to_tmp(f["bytes"], original_name=f["name"]) for f in fc1_list]
        try:
            df_1, col_map_1, err_1, warn_1 = _load_multi_files(paths_1, ["LCF","Creep"], "t1")
        except Exception as _ex:
            import traceback as _tb
            st.error(f"❌ File processing failed: {_ex}")
            st.code(_tb.format_exc())
            _show_file_diagnostics(paths_1[0], ["LCF","Creep"])
            err_1 = [str(_ex)]; warn_1 = []; df_1 = None; col_map_1 = {}
        if err_1:
            st.error("❌ " + "\n".join(err_1))
            _show_file_diagnostics(paths_1[0], ["LCF","Creep"])
        elif df_1 is None:
            pass
        else:
            for log_line in warn_1:
                if log_line.startswith("OK"):
                    st.success(log_line)
                elif "WARNING" in log_line:
                    st.warning(log_line)
                elif log_line.startswith("SKIPPED") or log_line.startswith("ERROR"):
                    st.warning(log_line)

            wdf_1, udf_1, types_1, years_1 = core.build_weekly(df_1, col_map_1, caps_1)
            annual_1 = get_annual_totals(wdf_1, types_1, years_1)

            # Wire data into the sidebar chat assistant
            st.session_state["data_tool1"] = {
                "types":      list(types_1),
                "years":      [int(y) for y in years_1],
                "annual":     {int(y): {t: float(annual_1.get(int(y),{}).get(t,0))
                               for t in types_1} for y in years_1},
                "util_df":    udf_1,
                "capacities": caps_1,
            }

            # KPI cards
            st.markdown('<div class="section-label">📊 Key Metrics — Latest Year</div>', unsafe_allow_html=True)
            show_kpi_row(types_1, annual_1, caps_1, years_1)

            # Charts in tabs
            tabs = st.tabs(["📈 Utilization Trend", "📊 YoY + Pie Charts",
                             "⚡ Capacity vs Demand", "🗓 Gantt", "📋 Data Table"])

            with tabs[0]:
                fig_util = chart_utilization_line(udf_1, types_1, years_1,
                                                   THEME_COLORS[theme_name])
                st.plotly_chart(fig_util, use_container_width=True)

            with tabs[1]:
                fig_bar_1, fig_pies_1 = chart_yoy_bar_and_pie(annual_1, types_1, years_1,
                                                             THEME_COLORS[theme_name])
                st.plotly_chart(fig_bar_1, use_container_width=True)
                st.plotly_chart(fig_pies_1, use_container_width=True)

            with tabs[2]:
                fig_cap = chart_capacity_bar(wdf_1, types_1, caps_1, years_1)
                st.plotly_chart(fig_cap, use_container_width=True)

            with tabs[3]:
                gantt_year = CURRENT_YEAR if CURRENT_YEAR in [int(y) for y in years_1] else int(max(years_1))
                gantt_week = CURRENT_WEEK if gantt_year == CURRENT_YEAR else 52
                fig_gantt = chart_gantt(wdf_1, types_1, caps_1, gantt_year, gantt_week)
                st.plotly_chart(fig_gantt, use_container_width=True)

            with tabs[4]:
                st.markdown("**Utilization Summary (Peak & Avg per year)**")
                show_util_table(udf_1, types_1, years_1)
                show_annual_summary_table(annual_1, types_1, years_1, caps_1)
                st.markdown("---")
                st.markdown("**✏️ Editable Data Export** — download a pre-filled Excel you can edit and re-upload")
                editable_bytes_1 = make_editable_excel(
                    wdf_1, udf_1, annual_1, types_1, caps_1, years_1,
                    title="LCF & Creep Lab"
                )
                st.download_button(
                    "⬇️ Download Editable Excel (Weekly Data + Summary)",
                    data=editable_bytes_1,
                    file_name="LCF_Creep_Editable.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t1_editable_dl",
                )

            # Generate Excel
            st.markdown("---")
            st.markdown('<div class="section-label">💾 Generate Full Excel Dashboard</div>', unsafe_allow_html=True)
            if st.button("⚡ Generate Full Excel Dashboard", key="t1_gen"):
                with st.spinner("Generating Excel dashboard..."):
                    out1, warns = gen_module.generate_lcf_creep(paths_1, caps_1, theme_1)
                    with open(out1, 'rb') as f: excel_bytes = f.read()
                st.download_button(
                    "⬇️ Download LCF_Creep_Dashboard.xlsx",
                    data=excel_bytes, file_name="LCF_Creep_Dashboard.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t1_dl",
                )
                st.success("✅ Dashboard ready! Click the button above to download.")
    else:
        st.info("👆 Upload one or more Excel files to begin. Files should contain LCF and/or Creep data — "
                "any layout works (long, wide, messy headers, or one monthly-block file per type).")


# ══════════════════════════════════════════════════════════════════
#  TOOL 2 — COATING LABS
# ══════════════════════════════════════════════════════════════════

elif "Tool 2" in tool:
    st.markdown("""
    <div class="tool-header" style="background: linear-gradient(90deg, #1F5C1A 0%, #70AD47 100%);">
        <h2>🟢 Coating Labs Dashboard</h2>
        <p>Cold Spray · HVOF · Plasma — individual + combined capacity tracking</p>
    </div>
    """, unsafe_allow_html=True)

    # ── File upload ───────────────────────────────────────
    st.markdown('<div class="section-label">📂 Upload Excel File(s)</div>', unsafe_allow_html=True)
    uploaded_2 = st.file_uploader(
        "", type=["xlsx", "xls"], key="t2_file",
        accept_multiple_files=True, label_visibility="collapsed",
        help="Upload one combined file, or one file per lab type (e.g. separate "
             "monthly-block files for Cold Spray, HVOF, Plasma) — they'll be merged.",
    )
    fc2_list = _persist_upload_multi("_fc_t2", uploaded_2 or [])
    if fc2_list and not uploaded_2:
        st.success(f"✅ {len(fc2_list)} file(s) loaded: {', '.join(f['name'] for f in fc2_list)}")
    elif uploaded_2:
        st.success(f"✅ {len(uploaded_2)} file(s) loaded: {', '.join(f.name for f in uploaded_2)}")
    if fc2_list:
        if st.button("✖ Remove file(s)", key="t2_remove"):
            st.session_state["_fc_t2"] = None
            st.rerun()

    # ── Capacity config ───────────────────────────────────
    st.markdown('<div class="section-label">⚙️ Capacity Configuration — Each Lab is Independent</div>', unsafe_allow_html=True)
    st.info("💡 Each lab has its **own independent capacity** of 350 samples/year by default. Edit any lab below.")

    c1, c2, c3 = st.columns(3)
    with c1:
        cs_cap = st.number_input("🔵 Cold Spray (samples/yr)",
                                  value=350, min_value=1, max_value=9999, key="t2_cs")
    with c2:
        hvof_cap = st.number_input("🟠 HVOF (samples/yr)",
                                    value=350, min_value=1, max_value=9999, key="t2_hvof")
    with c3:
        plasma_cap = st.number_input("🟣 Plasma (samples/yr)",
                                      value=350, min_value=1, max_value=9999, key="t2_pl")
    ind_caps_2 = {"Cold Spray": cs_cap, "HVOF": hvof_cap, "Plasma": plasma_cap}
    st.caption(f"Weekly caps → Cold Spray: **{cs_cap/52:.1f}**/wk  |  HVOF: **{hvof_cap/52:.1f}**/wk  |  Plasma: **{plasma_cap/52:.1f}**/wk")

    theme_name_2 = st.selectbox("Color Theme", list(core.COLOR_THEMES.keys()), key="t2_theme")
    theme_2 = core.COLOR_THEMES[theme_name_2]

    if fc2_list:
        COATING_TYPES = ["Cold Spray", "HVOF", "Plasma"]
        paths_2 = [_bytes_to_tmp(f["bytes"], original_name=f["name"]) for f in fc2_list]
        try:
            df_2, col_map_2, err_2, warn_2 = _load_multi_files(paths_2, COATING_TYPES, "t2")
        except Exception as _ex:
            import traceback as _tb
            st.error(f"❌ File processing failed: {_ex}")
            st.code(_tb.format_exc())
            _show_file_diagnostics(paths_2[0], COATING_TYPES)
            err_2 = [str(_ex)]; warn_2 = []; df_2 = None; col_map_2 = {}
        if err_2:
            st.error("❌ " + "\n".join(err_2))
            _show_file_diagnostics(paths_2[0], COATING_TYPES)
        elif df_2 is None:
            pass
        else:
            for log_line in warn_2:
                if log_line.startswith("OK"):
                    st.success(log_line)
                elif "WARNING" in log_line:
                    st.warning(log_line)
                elif log_line.startswith("SKIPPED") or log_line.startswith("ERROR"):
                    st.warning(log_line)

            wdf_2, udf_2, types_2, years_2 = core.build_weekly(df_2, col_map_2, ind_caps_2)
            annual_2 = get_annual_totals(wdf_2, types_2, years_2)

            # Wire data into the sidebar chat assistant
            st.session_state["data_tool2"] = {
                "types":      list(types_2),
                "years":      [int(y) for y in years_2],
                "annual":     {int(y): {t: float(annual_2.get(int(y),{}).get(t,0))
                               for t in types_2} for y in years_2},
                "util_df":    udf_2,
                "capacities": ind_caps_2,
            }

            # Combined utilization KPIs
            st.markdown('<div class="section-label">📊 Lab Utilization — Latest Year</div>', unsafe_allow_html=True)
            last_y2 = int(max(years_2))

            col_kpi1, col_kpi2, col_kpi3 = st.columns(3)
            with col_kpi1:
                combined_demand = sum(annual_2[last_y2].get(t, 0) for t in types_2)
                total_cap = sum(ind_caps_2.values())
                comb_util = combined_demand / total_cap if total_cap > 0 else 0
                comb_vacant = max(0, total_cap - combined_demand)
                cls = "red" if comb_util > 1 else "amber" if comb_util >= 0.8 else "green"
                vcol = "#888" if comb_vacant > 0 else "#FF4444"
                st.markdown(f"""<div class="metric-card {cls}">
                    <div style="font-size:11px;color:#666;font-weight:600;">Total Demand ({last_y2})</div>
                    <div style="font-size:28px;font-weight:800;color:#1F5C1A;">{combined_demand:.0f}</div>
                    <div style="font-size:13px;">across all 3 labs (cap: {total_cap})</div>
                    <div style="font-size:12px;color:{vcol};margin-top:2px;">🕳️ Vacant: <b>{comb_vacant:.0f}</b>/yr ({comb_vacant/52:.1f}/wk)</div>
                </div>""", unsafe_allow_html=True)
            with col_kpi2:
                st.markdown(f"""<div class="metric-card">
                    <div style="font-size:11px;color:#666;font-weight:600;">Weekly Demand Avg</div>
                    <div style="font-size:28px;font-weight:800;color:#1F5C1A;">{combined_demand/52:.1f}</div>
                    <div style="font-size:13px;">samples/week total</div>
                </div>""", unsafe_allow_html=True)
            with col_kpi3:
                peak_wk = sum(wdf_2[wdf_2["Year"]==last_y2][t].max() for t in types_2)
                total_wk_cap = sum(ind_caps_2.values()) / 52
                peak_util = peak_wk / total_wk_cap if total_wk_cap > 0 else 0
                cls3 = "red" if peak_util > 1 else "amber" if peak_util >= 0.8 else "green"
                st.markdown(f"""<div class="metric-card {cls3}">
                    <div style="font-size:11px;color:#666;font-weight:600;">Peak Week Demand ({last_y2})</div>
                    <div style="font-size:28px;font-weight:800;color:#1F5C1A;">{peak_wk:.1f}</div>
                    <div style="font-size:13px;">samples ({peak_util:.0%} of total weekly cap)</div>
                </div>""", unsafe_allow_html=True)

            show_kpi_row(types_2, annual_2, ind_caps_2, years_2)

            tabs2 = st.tabs(["📈 Utilization", "📊 YoY + Pie Charts",
                              "⚡ Capacity vs Demand", "🗓 Gantt", "📋 Data Table"])

            with tabs2[0]:
                fig_u2 = chart_utilization_line(udf_2, types_2, years_2,
                                                 THEME_COLORS[theme_name_2])
                st.plotly_chart(fig_u2, use_container_width=True)

            with tabs2[1]:
                fig_bar2, fig_pies_2 = chart_yoy_bar_and_pie(annual_2, types_2, years_2,
                                                               THEME_COLORS[theme_name_2])
                st.plotly_chart(fig_bar2, use_container_width=True)
                st.plotly_chart(fig_pies_2, use_container_width=True)

            with tabs2[2]:
                fig_cap2 = chart_capacity_bar(wdf_2, types_2, ind_caps_2, years_2)
                st.plotly_chart(fig_cap2, use_container_width=True)

            with tabs2[3]:
                gantt_year2 = CURRENT_YEAR if CURRENT_YEAR in [int(y) for y in years_2] else int(max(years_2))
                gantt_week2 = CURRENT_WEEK if gantt_year2 == CURRENT_YEAR else 52
                fig_g2 = chart_gantt(wdf_2, types_2, ind_caps_2, gantt_year2, gantt_week2)
                st.plotly_chart(fig_g2, use_container_width=True)

            with tabs2[4]:
                show_util_table(udf_2, types_2, years_2)
                show_annual_summary_table(annual_2, types_2, years_2, ind_caps_2)
                st.markdown("---")
                st.markdown("**✏️ Editable Data Export** — download a pre-filled Excel you can edit and re-upload")
                editable_bytes_2 = make_editable_excel(
                    wdf_2, udf_2, annual_2, types_2, ind_caps_2, years_2,
                    title="Coating Labs"
                )
                st.download_button(
                    "⬇️ Download Editable Excel (Weekly Data + Summary)",
                    data=editable_bytes_2,
                    file_name="Coating_Labs_Editable.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t2_editable_dl",
                )

            st.markdown("---")
            st.markdown('<div class="section-label">💾 Generate Full Excel Dashboard</div>', unsafe_allow_html=True)
            if st.button("⚡ Generate Full Excel Dashboard", key="t2_gen"):
                with st.spinner("Generating..."):
                    out2, warns2 = gen_module.generate_coating(paths_2, ind_caps_2, theme_2)
                    with open(out2, 'rb') as f: excel_bytes2 = f.read()
                st.download_button(
                    "⬇️ Download Coating_Labs_Dashboard.xlsx",
                    data=excel_bytes2, file_name="Coating_Labs_Dashboard.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t2_dl",
                )
                st.success("✅ Done!")
    else:
        st.info("👆 Upload one or more Excel files with Cold Spray, HVOF, and/or Plasma data — "
                "any layout works (long, wide, messy headers, or one monthly-block file per type).")


# ══════════════════════════════════════════════════════════════════
#  TOOL 3 — THERMAL LAB
# ══════════════════════════════════════════════════════════════════

elif "Tool 3" in tool:
    st.markdown("""
    <div class="tool-header" style="background: linear-gradient(90deg, #7F3F00 0%, #C55A11 100%);">
        <h2>🟠 Tool 3 — Thermal Lab Dashboard</h2>
        <p>Thermal Rig occupancy planning — per-rig capacity tracking</p>
    </div>
    """, unsafe_allow_html=True)

    # ── File upload ───────────────────────────────────────
    st.markdown('<div class="section-label">📂 Upload Excel File</div>', unsafe_allow_html=True)
    uploaded_t3 = st.file_uploader("", type=["xlsx","xls"], key="t3_file",
                                    label_visibility="collapsed")
    fc3 = _persist_upload("_fc_t3", uploaded_t3)
    if fc3 and uploaded_t3 is None:
        st.success(f"✅ {fc3['name']} loaded")
    if fc3:
        if st.button("✖ Remove file", key="t3_remove"):
            st.session_state["_fc_t3"] = None
            st.rerun()

    # ── Capacity config ───────────────────────────────────
    st.markdown('<div class="section-label">⚙️ Rig Capacity (samples/year)</div>', unsafe_allow_html=True)

    r1_cap = st.number_input("🔶 Thermal Rig capacity", value=200, min_value=1, max_value=9999,
                              key="t3_r1")
    st.caption(f"→ {r1_cap/52:.1f} samples/wk  |  Vacant if below capacity")

    rig_caps_t3 = {"Thermal Rig": r1_cap}
    thermal_types = ["Thermal Rig"]

    theme_name_t3 = st.selectbox("Color Theme", list(core.COLOR_THEMES.keys()), key="t3_theme")
    theme_t3 = core.COLOR_THEMES[theme_name_t3]

    if fc3:
        paths_t3 = [_bytes_to_tmp(fc3["bytes"])]
        try:
            df_t3, col_map_t3, err_t3, warn_t3 = core.load_and_filter(paths_t3[0], thermal_types)
        except Exception as _ex:
            import traceback as _tb
            st.error(f"❌ File processing failed: {_ex}")
            st.code(_tb.format_exc())
            _show_file_diagnostics(paths_t3[0], thermal_types)
            err_t3 = [str(_ex)]; warn_t3 = []; df_t3 = None; col_map_t3 = {}
        if err_t3:
            st.error("❌ " + "\n".join(err_t3))
            _show_file_diagnostics(paths_t3[0], thermal_types)
        elif df_t3 is None:
            pass
        else:
            if warn_t3:
                for w in warn_t3:
                    st.warning(w)

            wdf_t3, udf_t3, types_t3, years_t3 = core.build_weekly(df_t3, col_map_t3, rig_caps_t3)
            annual_t3 = get_annual_totals(wdf_t3, types_t3, years_t3)

            # Wire data into the sidebar chat assistant
            st.session_state["data_tool3"] = {
                "types":      list(types_t3),
                "years":      [int(y) for y in years_t3],
                "annual":     {int(y): {t: float(annual_t3.get(int(y),{}).get(t,0))
                               for t in types_t3} for y in years_t3},
                "util_df":    udf_t3,
                "capacities": rig_caps_t3,
            }

            # KPI cards
            st.markdown('<div class="section-label">📊 Key Metrics — Latest Year</div>',
                        unsafe_allow_html=True)
            show_kpi_row(types_t3, annual_t3, rig_caps_t3, years_t3)

            # Charts in tabs
            tabs_t3 = st.tabs(["📈 Utilization Trend", "📊 YoY + Pie Charts",
                                "⚡ Capacity vs Demand", "🗓 Gantt", "📋 Data Table"])

            with tabs_t3[0]:
                fig_ut3 = chart_utilization_line(udf_t3, types_t3, years_t3,
                                                  THEME_COLORS[theme_name_t3])
                st.plotly_chart(fig_ut3, use_container_width=True)

            with tabs_t3[1]:
                fig_bar_t3, fig_pies_t3 = chart_yoy_bar_and_pie(
                    annual_t3, types_t3, years_t3, THEME_COLORS[theme_name_t3])
                st.plotly_chart(fig_bar_t3, use_container_width=True)
                st.plotly_chart(fig_pies_t3, use_container_width=True)

            with tabs_t3[2]:
                fig_cap_t3 = chart_capacity_bar(wdf_t3, types_t3, rig_caps_t3, years_t3)
                st.plotly_chart(fig_cap_t3, use_container_width=True)

            with tabs_t3[3]:
                gantt_yr_t3 = (CURRENT_YEAR if CURRENT_YEAR in [int(y) for y in years_t3]
                               else int(max(years_t3)))
                gantt_wk_t3 = CURRENT_WEEK if gantt_yr_t3 == CURRENT_YEAR else 52
                fig_gantt_t3 = chart_gantt(wdf_t3, types_t3, rig_caps_t3,
                                            gantt_yr_t3, gantt_wk_t3)
                st.plotly_chart(fig_gantt_t3, use_container_width=True)

            with tabs_t3[4]:
                st.markdown("**Utilization Summary (Peak & Avg per year)**")
                show_util_table(udf_t3, types_t3, years_t3)
                show_annual_summary_table(annual_t3, types_t3, years_t3, rig_caps_t3)
                st.markdown("---")
                st.markdown("**✏️ Editable Data Export** — download a pre-filled Excel you can edit and re-upload")
                editable_bytes_t3 = make_editable_excel(
                    wdf_t3, udf_t3, annual_t3, types_t3, rig_caps_t3, years_t3,
                    title="Thermal Lab"
                )
                st.download_button(
                    "⬇️ Download Editable Excel (Weekly Data + Summary)",
                    data=editable_bytes_t3,
                    file_name="Thermal_Lab_Editable.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t3_editable_dl",
                )

            # Generate Full Excel Dashboard
            st.markdown("---")
            st.markdown('<div class="section-label">💾 Generate Full Excel Dashboard</div>',
                        unsafe_allow_html=True)
            if st.button("⚡ Generate Full Excel Dashboard", key="t3_gen"):
                with st.spinner("Generating Thermal Lab dashboard..."):
                    out_t3, warns_t3 = gen_module.generate_thermal(
                        paths_t3[0], rig_caps_t3, theme_t3)
                    with open(out_t3, 'rb') as f:
                        excel_bytes_t3 = f.read()
                st.download_button(
                    "⬇️ Download Thermal_Lab_Dashboard.xlsx",
                    data=excel_bytes_t3,
                    file_name="Thermal_Lab_Dashboard.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="t3_dl",
                )
                st.success("✅ Dashboard ready! Click the button above to download.")
    else:
        st.info("👆 Upload an Excel file. Columns: Year | Type (Thermal Rig) | Value  "
                "— or wide format: Year | Week | Thermal Rig")


# ══════════════════════════════════════════════════════════════════
#  TOOL 3 — COMPARISON
# ══════════════════════════════════════════════════════════════════

elif "Tool 4" in tool:
    st.markdown("""
    <div class="tool-header" style="background: linear-gradient(90deg, #3D1F6E 0%, #7B5EA7 100%);">
        <h2>🔴 Tool 4 — Comparison & PPT</h2>
        <p>Mechanical · Coating · Thermal — side-by-side comparison + PowerPoint export</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Multi-file upload ─────────────────────────────────
    st.markdown('<div class="section-label">📂 Upload Files (1–3 files, one per group or one combined)</div>',
                unsafe_allow_html=True)
    uploaded_3 = st.file_uploader(
        "",
        type=["xlsx","xls"],
        accept_multiple_files=True,
        key="t4_files",
        label_visibility="collapsed",
        help="Upload one combined file OR up to 3 separate files (one per lab group). Duplicates are merged.",
    )
    fc4_list = _persist_upload_multi("_fc_t4", uploaded_3 or [])
    if fc4_list and not uploaded_3:
        st.success(f"✅ {len(fc4_list)} file(s) loaded: {', '.join(f['name'] for f in fc4_list)}")
    elif uploaded_3:
        st.success(f"✅ {len(uploaded_3)} file(s) loaded: {', '.join(f.name for f in uploaded_3)}")
    if fc4_list:
        if st.button("✖ Remove files", key="t4_remove"):
            st.session_state["_fc_t4"] = None
            st.rerun()

    # ── Capacities — 3 columns, one per group ─────────────
    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        st.markdown('<div class="section-label">🔵 Mechanical Capacities</div>', unsafe_allow_html=True)
        mc1, mc2 = st.columns(2)
        with mc1:
            lcf_cap3  = st.number_input("LCF (samples/yr)",  value=50, min_value=1, key="t4_lcf")
        with mc2:
            creep_cap3 = st.number_input("Creep (samples/yr)", value=22, min_value=1, key="t4_creep")
        caps_a3 = {"LCF": lcf_cap3, "Creep": creep_cap3}
        st.caption(f"Combined: **{sum(caps_a3.values())}**/yr")

    with cc2:
        st.markdown('<div class="section-label">🟢 Coating Capacities</div>', unsafe_allow_html=True)
        bc1, bc2, bc3 = st.columns(3)
        with bc1: cs3 = st.number_input("Cold Spray", value=350, min_value=1, key="t4_cs")
        with bc2: hv3 = st.number_input("HVOF",       value=350, min_value=1, key="t4_hv")
        with bc3: pl3 = st.number_input("Plasma",     value=350, min_value=1, key="t4_pl")
        caps_b3 = {"Cold Spray": cs3, "HVOF": hv3, "Plasma": pl3}
        st.caption(f"CS {cs3/52:.1f} | HVOF {hv3/52:.1f} | Plasma {pl3/52:.1f}  /wk")

    with cc3:
        st.markdown('<div class="section-label">🟠 Thermal Capacities</div>', unsafe_allow_html=True)
        tr1 = st.number_input("Thermal Rig", value=200, min_value=1, key="t4_r1")
        caps_c3 = {"Thermal Rig": tr1}
        st.caption(f"→ {tr1/52:.1f}/wk")

    theme_name_3 = st.selectbox("Color Theme", list(core.COLOR_THEMES.keys()), key="t4_theme")
    theme_3 = core.COLOR_THEMES[theme_name_3]

    if fc4_list:
        paths_3 = [_bytes_to_tmp(f["bytes"], original_name=f["name"]) for f in fc4_list]
        merged_df, col_map_3, errors_3, file_log_3 = _load_multi_files(paths_3, ALL_TYPES, "t4")

        for log_line in file_log_3:
            if log_line.startswith("OK"):
                st.success(log_line)
            elif "WARNING" in log_line:
                st.warning(log_line)
            elif log_line.startswith("SKIPPED") or log_line.startswith("ERROR"):
                st.warning(log_line)

        if errors_3:
            st.error("❌ " + "\n".join(errors_3))
        else:
            tc3_col_name = col_map_3["type"]
            df_a3 = merged_df[merged_df[tc3_col_name].isin(GROUP_A["types"])].copy()
            df_b3 = merged_df[merged_df[tc3_col_name].isin(GROUP_B["types"])].copy()
            df_c3 = merged_df[merged_df[tc3_col_name].isin(list(caps_c3.keys()))].copy()

            if df_a3.empty:
                st.error(f"No Mechanical lab data found. Need: {GROUP_A['types']}")
                st.stop()
            if df_b3.empty:
                st.error(f"No Coating lab data found. Need: {GROUP_B['types']}")
                st.stop()

            has_c3 = not df_c3.empty
            if not has_c3:
                st.warning("⚠️ No Thermal Rig data found — Group C omitted from charts.")

            wdf_a3, udf_a3, types_a3, years_a3 = core.build_weekly(df_a3, col_map_3, caps_a3)
            wdf_b3, udf_b3, types_b3, years_b3 = core.build_weekly(df_b3, col_map_3, caps_b3)
            years_3 = sorted(set(list(years_a3)) | set(list(years_b3)))
            annual_a3 = get_annual_totals(wdf_a3, types_a3, years_a3)
            annual_b3 = get_annual_totals(wdf_b3, types_b3, years_b3)

            if has_c3:
                wdf_c3, udf_c3, types_c3, years_c3 = core.build_weekly(df_c3, col_map_3, caps_c3)
                years_3 = sorted(set(years_3) | set(list(years_c3)))
                annual_c3 = get_annual_totals(wdf_c3, types_c3, years_c3)
            else:
                wdf_c3 = udf_c3 = types_c3 = years_c3 = None
                annual_c3 = {}

            # ── KPI summary ───────────────────────────────
            last_y3 = int(max(years_3))
            sum_caps_a3 = sum(caps_a3.values())
            sum_caps_b3 = sum(caps_b3.values())
            sum_caps_c3 = sum(caps_c3.values())

            dem_a3 = sum(annual_a3.get(last_y3, {}).get(t, 0) for t in types_a3)
            dem_b3 = sum(annual_b3.get(last_y3, {}).get(t, 0) for t in types_b3)
            dem_c3 = (sum(annual_c3.get(last_y3, {}).get(t, 0) for t in types_c3)
                      if has_c3 else 0)
            u_a3 = dem_a3 / sum_caps_a3 if sum_caps_a3 > 0 else 0
            u_b3 = dem_b3 / sum_caps_b3 if sum_caps_b3 > 0 else 0
            u_c3 = dem_c3 / sum_caps_c3 if sum_caps_c3 > 0 else 0

            st.markdown('<div class="section-label">📊 Group Utilization — Latest Year</div>',
                        unsafe_allow_html=True)
            kpi_col1, kpi_col2, kpi_col3 = st.columns(3)
            for col_kpi, label, dem, cap_sum, util, color in [
                (kpi_col1, "🔵 Mechanical", dem_a3, sum_caps_a3, u_a3, "#1A4E8A"),
                (kpi_col2, "🟢 Coating",    dem_b3, sum_caps_b3, u_b3, "#1F5C1A"),
                (kpi_col3, "🟠 Thermal",    dem_c3, sum_caps_c3, u_c3, "#7F3F00"),
            ]:
                cls = "red" if util > 1 else "amber" if util >= 0.8 else "green"
                with col_kpi:
                    st.markdown(f"""<div class="metric-card {cls}">
                        <div style="font-size:12px;color:#666;font-weight:700;">{label} ({last_y3})</div>
                        <div style="font-size:26px;font-weight:800;color:{color};">{dem:.0f} samples</div>
                        <div style="font-size:13px;">Utilization: {util:.0%} of {cap_sum}/yr</div>
                    </div>""", unsafe_allow_html=True)

            # ── Build figures ─────────────────────────────
            # Tab 1 — 3-panel comparison bar chart
            def chart_comparison_3groups(annual_a, annual_b, annual_c, types_a, types_b, types_c,
                                          years, cap_a, cap_b, cap_c, has_c):
                n_panels = 3 if has_c else 2
                panel_titles = [GROUP_A["name"], GROUP_B["name"]]
                if has_c:
                    panel_titles.append("Thermal Lab")
                fig = make_subplots(rows=1, cols=n_panels, subplot_titles=panel_titles)
                col_maps_a = ["#1F3864","#2E75B6","#9DC3E6"]
                col_maps_b = ["#1F5C1A","#70AD47","#C6EFCE"]
                col_maps_c = ["#7F3F00","#C55A11","#F4B183"]
                for ti, t in enumerate(types_a):
                    vals = [annual_a.get(int(y), {}).get(t, 0) for y in years]
                    fig.add_trace(go.Bar(name=t, x=[str(y) for y in years], y=vals,
                                         marker_color=col_maps_a[ti % 3],
                                         text=[f"{v:.0f}" for v in vals],
                                         textposition="outside", legendgroup="A"), row=1, col=1)
                for ti, t in enumerate(types_b):
                    vals = [annual_b.get(int(y), {}).get(t, 0) for y in years]
                    fig.add_trace(go.Bar(name=t, x=[str(y) for y in years], y=vals,
                                         marker_color=col_maps_b[ti % 3],
                                         text=[f"{v:.0f}" for v in vals],
                                         textposition="outside", legendgroup="B"), row=1, col=2)
                if has_c:
                    for ti, t in enumerate(types_c):
                        vals = [annual_c.get(int(y), {}).get(t, 0) for y in years]
                        fig.add_trace(go.Bar(name=t, x=[str(y) for y in years], y=vals,
                                             marker_color=col_maps_c[ti % 3],
                                             text=[f"{v:.0f}" for v in vals],
                                             textposition="outside", legendgroup="C"), row=1, col=3)
                # Capacity reference lines + corner badges.
                # The old approach attached the "Cap A:72" text directly to
                # the dashed line via annotation_text, which Plotly places
                # at the line's height — right where tall bars and their
                # "outside" value labels also sit, causing the text to
                # collide/overlap (e.g. "Cap A:72" over a "66" bar label).
                # Pinning the badge to a fixed corner of each subplot
                # (paper-domain coordinates, not data coordinates) keeps it
                # legible regardless of bar heights.
                def _cap_badge(row, col, cap_val, color, label):
                    fig.add_hline(y=cap_val, line_dash="dot", line_color=color,
                                   row=row, col=col)
                    fig.add_annotation(
                        text=f"{label}: {cap_val:.0f}/yr",
                        xref="x domain", yref="y domain",
                        x=0.02, y=0.97, xanchor="left", yanchor="top",
                        showarrow=False,
                        font=dict(size=10, color=color),
                        bgcolor="rgba(255,255,255,0.88)",
                        bordercolor=color, borderwidth=1, borderpad=3,
                        row=row, col=col,
                    )

                _cap_badge(1, 1, cap_a, "#1A4E8A", "Cap A")
                _cap_badge(1, 2, cap_b, "#1F5C1A", "Cap B")
                if has_c:
                    _cap_badge(1, 3, cap_c, "#7F3F00", "Cap C")
                fig.update_layout(barmode="group", height=480,
                                   title="Annual Demand Comparison — All 3 Lab Groups",
                                   font=dict(family="Arial", size=11),
                                   paper_bgcolor="white", plot_bgcolor="white",
                                   legend=dict(orientation="h", y=-0.2),
                                   margin=dict(b=100, t=60))
                return fig

            # Tab 3 — 3-group utilisation line chart
            def chart_util_3groups(wdf_a, wdf_b, wdf_c, types_a, types_b, types_c,
                                    years, cap_a, cap_b, cap_c, has_c):
                months = ["Jan","Feb","Mar","Apr","May","Jun",
                          "Jul","Aug","Sep","Oct","Nov","Dec"]
                fig = go.Figure()
                plot_yrs = list(years)[-3:]
                dash_styles = ["solid", "dash", "dot"]
                for yi, year in enumerate(plot_yrs):
                    dash = dash_styles[yi % 3]
                    for wdf, cap_total, types, label, color in [
                        (wdf_a, cap_a, types_a, "Mechanical", "#1A4E8A"),
                        (wdf_b, cap_b, types_b, "Coating",    "#1F5C1A"),
                    ] + ([(wdf_c, cap_c, types_c, "Thermal", "#C55A11")] if has_c else []):
                        yd = wdf[wdf["Year"] == year].copy()
                        yd["_m"] = yd["Week"].apply(core.week_to_month)
                        monthly = []
                        for m in range(12):
                            tot = sum(yd[yd["_m"]==m+1][t].mean()
                                      for t in types if t in yd.columns)
                            cap_wk = cap_total / 52
                            util = tot / cap_wk if cap_wk > 0 and not math.isnan(tot) else 0
                            monthly.append(util)
                        fig.add_trace(go.Scatter(
                            x=months, y=monthly,
                            name=f"{label} {year}",
                            line=dict(color=color, width=2, dash=dash),
                            mode="lines+markers", marker=dict(size=5),
                        ))
                fig.add_hline(y=1.0, line_dash="dot", line_color="#FF4444",
                               annotation_text="100%", annotation_position="right")
                fig.add_hline(y=0.8, line_dash="dot", line_color="#FFD700",
                               annotation_text="80%", annotation_position="right")
                fig.update_layout(
                    title="Monthly Utilization % — All Groups (last 3 years)",
                    xaxis_title="Month", yaxis_title="Utilization",
                    yaxis_tickformat=".0%", height=440,
                    font=dict(family="Arial", size=12),
                    paper_bgcolor="white", plot_bgcolor="white",
                    legend=dict(orientation="h", y=-0.4), margin=dict(b=120),
                )
                return fig

            fig_cmp3 = chart_comparison_3groups(
                annual_a3, annual_b3, annual_c3,
                types_a3, types_b3, types_c3 or [],
                years_3, sum_caps_a3, sum_caps_b3, sum_caps_c3, has_c3)

            fig_util3 = chart_util_3groups(
                wdf_a3, wdf_b3, wdf_c3, types_a3, types_b3, types_c3 or [],
                years_3, sum_caps_a3, sum_caps_b3, sum_caps_c3, has_c3)

            # Gantt figures
            g_yr3 = CURRENT_YEAR if CURRENT_YEAR in [int(y) for y in years_3] else int(max(years_3))
            g_wk3 = CURRENT_WEEK if g_yr3 == CURRENT_YEAR else 52
            fig_ga3 = chart_gantt(wdf_a3, types_a3, caps_a3, g_yr3, g_wk3)
            fig_gb3 = chart_gantt(wdf_b3, types_b3, caps_b3, g_yr3, g_wk3)
            fig_gc3 = (chart_gantt(wdf_c3, types_c3, caps_c3, g_yr3, g_wk3)
                       if has_c3 else None)

            # ── TABS ─────────────────────────────────────
            tabs4 = st.tabs(["📊 Comparison Charts", "🥧 YoY Pie Charts",
                              "📈 Utilization Trend", "🗓 Gantt", "📋 Data Tables"])

            with tabs4[0]:
                st.plotly_chart(fig_cmp3, use_container_width=True)

            with tabs4[1]:
                st.subheader("🔵 Mechanical Labs — Process Share")
                _, fig_pies_a3 = chart_yoy_bar_and_pie(annual_a3, types_a3, years_a3,
                                                         THEME_COLORS[theme_name_3])
                st.plotly_chart(fig_pies_a3, use_container_width=True)

                st.subheader("🟢 Coating Labs — Process Share")
                _, fig_pies_b3 = chart_yoy_bar_and_pie(annual_b3, types_b3, years_b3,
                                                         THEME_COLORS[theme_name_3])
                st.plotly_chart(fig_pies_b3, use_container_width=True)

                if has_c3:
                    st.subheader("🟠 Thermal Lab — Process Share")
                    _, fig_pies_c3 = chart_yoy_bar_and_pie(annual_c3, types_c3, years_c3,
                                                             THEME_COLORS[theme_name_3])
                    st.plotly_chart(fig_pies_c3, use_container_width=True)

            with tabs4[2]:
                st.plotly_chart(fig_util3, use_container_width=True)

            with tabs4[3]:
                gcol1, gcol2, gcol3 = st.columns(3)
                with gcol1:
                    st.markdown("**🔵 Mechanical**")
                    st.plotly_chart(fig_ga3, use_container_width=True)
                with gcol2:
                    st.markdown("**🟢 Coating**")
                    st.plotly_chart(fig_gb3, use_container_width=True)
                with gcol3:
                    st.markdown("**🟠 Thermal**")
                    if has_c3:
                        st.plotly_chart(fig_gc3, use_container_width=True)
                    else:
                        st.info("No Thermal data uploaded.")

            with tabs4[4]:
                st.markdown("**🔵 Mechanical Labs**")
                show_util_table(udf_a3, types_a3, years_a3)
                st.markdown("**🟢 Coating Labs**")
                show_util_table(udf_b3, types_b3, years_b3)
                if has_c3:
                    st.markdown("**🟠 Thermal Lab**")
                    show_util_table(udf_c3, types_c3, years_c3)

            # ── Export buttons ────────────────────────────
            st.markdown("---")
            st.markdown('<div class="section-label">💾 Export Dashboards</div>',
                        unsafe_allow_html=True)
            exp_col1, exp_col2 = st.columns(2)

            with exp_col1:
                if st.button("⚡ Generate Full Excel Dashboard", key="t4_gen"):
                    with st.spinner("Generating Excel..."):
                        out3, warns3 = gen_module.generate_comparison(
                            paths_3, caps_a3, caps_b3, theme_3, caps_c=caps_c3)
                        with open(out3, 'rb') as f:
                            excel_bytes3 = f.read()
                    st.download_button(
                        "⬇️ Download Lab_Comparison_Dashboard.xlsx",
                        data=excel_bytes3,
                        file_name="Lab_Comparison_Dashboard.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="t4_dl",
                    )
                    st.success("✅ Excel ready!")

            with exp_col2:
                if st.button("📊 Generate PPT Report", key="t4_ppt_gen"):
                    with st.spinner("Building PowerPoint (rendering charts to PNG)…"):
                        try:
                            ppt_bytes = gen_module.generate_comparison_ppt(
                                wdf_a3, wdf_b3, wdf_c3,
                                annual_a3, annual_b3, annual_c3,
                                types_a3, types_b3, types_c3 or [],
                                caps_a3, caps_b3, caps_c3,
                                years_3,
                                fig_cmp3, fig_util3,
                                [fig_ga3, fig_gb3, fig_gc3],
                            )
                            st.download_button(
                                "⬇️ Download Lab_Occupancy_Report.pptx",
                                data=ppt_bytes,
                                file_name="Lab_Occupancy_Report.pptx",
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                key="t4_ppt_dl",
                            )
                            st.success("✅ PowerPoint ready! 6 slides generated.")
                        except ImportError as e:
                            st.error(f"❌ {e}")
                        except Exception as e:
                            st.error(f"❌ PPT error: {str(e)[:200]}")
    else:
        st.info("👆 Upload 1–3 Excel files — one combined file or one per lab group (Mechanical, Coating, Thermal).")

    # ── Export section always visible in Tool 4 ───────────
    st.markdown("---")
    st.markdown('<div class="section-label">💾 Export Dashboards</div>', unsafe_allow_html=True)
    if not fc4_list:
        st.info("📤 Upload files above to enable Excel and PowerPoint export.")
    else:
        st.markdown("_Files loaded — click a button below to generate your export._")
