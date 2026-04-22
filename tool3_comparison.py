"""
Tool 3 — Lab Comparison Dashboard  (Multi-File Edition)
=========================================================
Supports ONE or MULTIPLE Excel files — data is merged automatically.
  Mode A: Single file with all lab types
  Mode B: Two files (one per group)
  Mode C: Multiple files — all merged by lab type + year

Run: python tool3_comparison.py
"""

import os, sys, threading, math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import lab_core as core

TOOL_TITLE  = "Lab Comparison Dashboard — Mechanical vs Coating Labs"
OUTPUT_FILE = "Lab_Comparison_Dashboard.xlsx"

GROUP_A = {"name":"Mechanical Labs","types":["LCF","Creep"],
           "caps":{"LCF":50,"Creep":22},"color":"1A4E8A","cap_total":72}
GROUP_B = {"name":"Coating Labs","types":["Cold Spray","HVOF","Plasma"],
           "caps":{"Cold Spray":140,"HVOF":120,"Plasma":90},"color":"1F5C1A","cap_total":350}
ALL_TYPES = GROUP_A["types"] + GROUP_B["types"]


# ────────────────────────────────────────────────────────────
#  MULTI-FILE MERGE ENGINE
# ────────────────────────────────────────────────────────────

def merge_files(file_paths):
    """
    Load >=1 Excel files, normalise columns, merge into single DataFrame.
    Duplicate Year+Type rows across files are SUMMED.
    Returns: (merged_df, col_map, errors, file_log)
    """
    frames = []
    errors = []
    file_log = []   # one line per file for the Data_Sources sheet

    for path in file_paths:
        fname = os.path.basename(path)
        try:
            raw = pd.read_excel(path)
        except Exception as e:
            errors.append(f"Cannot open '{fname}': {e}")
            continue

        if raw.empty:
            file_log.append(f"SKIPPED (empty): {fname}")
            continue

        col_map = core.detect_columns(raw)
        missing_cols = [k for k in ["year","type","value"] if k not in col_map]
        if missing_cols:
            errors.append(
                f"'{fname}' missing required columns: {missing_cols}. "
                f"Found: {list(raw.columns)}"
            )
            continue

        yc, tc, vc = col_map["year"], col_map["type"], col_map["value"]
        df = raw[[yc, tc, vc]].copy()
        df.columns = ["Year", "Type", "Value"]
        df["Value"] = pd.to_numeric(df["Value"], errors="coerce").fillna(0)
        df["Year"]  = pd.to_numeric(df["Year"],  errors="coerce")
        df = df.dropna(subset=["Year","Type","Value"])
        df["Year"]  = df["Year"].astype(int)
        df["Type"]  = df["Type"].astype(str).str.strip()

        # Normalise type names (case-insensitive)
        known = {t.lower(): t for t in ALL_TYPES}
        df["Type"] = df["Type"].apply(lambda x: known.get(x.lower(), x))

        rows_kept = len(df)
        frames.append(df)
        file_log.append(f"OK ({rows_kept} rows): {fname}")

    if errors:
        return None, {}, errors, file_log

    if not frames:
        errors.append("No valid data found in any file.")
        return None, {}, errors, file_log

    merged = pd.concat(frames, ignore_index=True)
    # Sum duplicates (same Year+Type from multiple files)
    merged = merged.groupby(["Year","Type"], as_index=False)["Value"].sum()

    found   = set(merged["Type"].unique())
    missing = set(ALL_TYPES) - found
    if missing:
        file_log.append(f"WARNING: Lab types not found in any file: {sorted(missing)}")

    col_map_out = {"year":"Year","type":"Type","value":"Value"}
    return merged, col_map_out, [], file_log


# ────────────────────────────────────────────────────────────
#  SHEET WRITERS
# ────────────────────────────────────────────────────────────

def write_data_sources(wb, file_paths, file_log):
    ws = wb.create_sheet("Data_Sources")
    core.banner(ws, 1, "Data Sources Loaded", cols=12)
    for c,h in enumerate(["#","File Name","Status","Full Path"],1):
        core.hdr(ws.cell(3,c,h), bg="1F3864", sz=9)
    core.cw(ws,1,5); core.cw(ws,2,35); core.cw(ws,3,18); core.cw(ws,4,60)
    for i,path in enumerate(file_paths,4):
        log = file_log[i-4] if i-4 < len(file_log) else ""
        status = "✓ OK" if log.startswith("OK") else "⚠ WARNING" if "WARNING" in log else "✗ ERROR"
        ws.cell(i,1,i-3).border=core.bdr()
        ws.cell(i,1).font=Font(name="Arial",size=9)
        c2=ws.cell(i,2,os.path.basename(path))
        c2.font=Font(name="Arial",size=9,bold=True); c2.border=core.bdr()
        c2.alignment=Alignment(horizontal="left")
        c3=ws.cell(i,3,status)
        c3.font=Font(name="Arial",size=9); c3.border=core.bdr()
        c3.fill=PatternFill("solid",start_color="C6EFCE" if "OK" in status else "FFD700")
        c3.alignment=Alignment(horizontal="center")
        c4=ws.cell(i,4,path)
        c4.font=Font(name="Arial",size=8,color="666666"); c4.border=core.bdr()
        c4.alignment=Alignment(horizontal="left")


def write_overview(wb, wdf_a, wdf_b, years):
    ws = wb.create_sheet("Overview")
    core.banner(ws, 1, f"Lab Comparison — Overview", cols=24, sz=13)
    R = 3
    for grp,wdf in [(GROUP_A,wdf_a),(GROUP_B,wdf_b)]:
        ws.merge_cells(start_row=R,start_column=1,end_row=R,end_column=3+len(years))
        c=ws.cell(R,1,f"  {grp['name']}  —  Capacity: {grp['cap_total']} samples/yr")
        c.font=Font(bold=True,size=11,name="Arial",color="FFFFFF")
        c.fill=PatternFill("solid",start_color=grp["color"])
        c.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[R].height=20; R+=1
        tgt_bg="2E75B6" if grp==GROUP_A else "375623"
        for ci,h in enumerate(["Lab Type"]+[str(y) for y in years]+["Total"],1):
            core.hdr(ws.cell(R,ci,h),bg=tgt_bg,sz=9); core.cw(ws,ci,14)
        R+=1
        for t in grp["types"]:
            ws.cell(R,1,t).font=Font(name="Arial",size=9,bold=True)
            ws.cell(R,1).border=core.bdr()
            ws.cell(R,1).alignment=Alignment(horizontal="left")
            row_total=0
            for ci,y in enumerate(years,2):
                val=wdf[wdf["Year"]==y][t].sum() if t in wdf.columns else 0
                cell=ws.cell(R,ci,round(val,1)); core.fmt(cell,number_format="#,##0.0")
                row_total+=val
            tc=ws.cell(R,2+len(years),round(row_total,1))
            core.fmt(tc,number_format="#,##0.0",bold=True)
            tc.fill=PatternFill("solid",start_color="BDD7EE" if grp==GROUP_A else "C6EFCE")
            R+=1
        ws.cell(R,1,"GROUP TOTAL")
        ws.cell(R,1).fill=PatternFill("solid",start_color=grp["color"])
        ws.cell(R,1).font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
        ws.cell(R,1).border=core.bdr()
        ws.cell(R,1).alignment=Alignment(horizontal="left")
        grand=0
        for ci,y in enumerate(years,2):
            yr_tot=sum(wdf[wdf["Year"]==y][t].sum() for t in grp["types"] if t in wdf.columns)
            cell=ws.cell(R,ci,round(yr_tot,1)); core.fmt(cell,number_format="#,##0.0",bold=True)
            util=yr_tot/grp["cap_total"] if grp["cap_total"]>0 else 0
            cell.fill=PatternFill("solid",start_color="FF4444" if util>1 else "FFD700" if util>=0.8 else "C6EFCE")
            if util>1: cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
            grand+=yr_tot
        tc=ws.cell(R,2+len(years),round(grand,1)); core.fmt(tc,number_format="#,##0.0",bold=True)
        tc.fill=PatternFill("solid",start_color=grp["color"])
        tc.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); R+=3
    # Util summary
    core.section_hdr(ws,R,"Annual Utilization vs Capacity",3+len(years)); R+=1
    for ci,h in enumerate(["Lab Group","Cap/yr"]+[str(y) for y in years],1):
        core.hdr(ws.cell(R,ci,h),bg="C00000",sz=9); core.cw(ws,ci,16)
    R+=1
    for grp,wdf in [(GROUP_A,wdf_a),(GROUP_B,wdf_b)]:
        ws.cell(R,1,grp["name"])
        ws.cell(R,1).fill=PatternFill("solid",start_color=grp["color"])
        ws.cell(R,1).font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
        ws.cell(R,1).border=core.bdr(); ws.cell(R,1).alignment=Alignment(horizontal="left")
        ws.cell(R,2,grp["cap_total"]); core.fmt(ws.cell(R,2),number_format="#,##0")
        for ci,y in enumerate(years,3):
            tot=sum(wdf[wdf["Year"]==y][t].sum() for t in grp["types"] if t in wdf.columns)
            util=tot/grp["cap_total"] if grp["cap_total"]>0 else 0
            cell=ws.cell(R,ci,round(util,4)); core.fmt(cell,number_format="0.0%")
            if util>1.0: cell.fill=PatternFill("solid",start_color="FF4444"); cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
            elif util>=0.8: cell.fill=PatternFill("solid",start_color="FFD700"); cell.font=Font(name="Arial",size=9,bold=True)
            else: cell.fill=PatternFill("solid",start_color="70AD47"); cell.font=Font(name="Arial",size=9,color="FFFFFF")
        R+=1


def write_comparison_chart(wb, wdf_a, wdf_b, years):
    ws = wb.create_sheet("Chart_Comparison")
    core.banner(ws,1,"Group Comparison — Annual Demand vs Capacity",cols=24)
    months=["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    HR=3; D0=4
    for ci,h in enumerate(["Year",f"{GROUP_A['name']} Demand",f"Cap({GROUP_A['cap_total']})",
                            f"{GROUP_B['name']} Demand",f"Cap({GROUP_B['cap_total']})"],1):
        core.hdr(ws.cell(HR,ci,h),bg="1F3864",sz=9); core.cw(ws,ci,22)
    DE=D0+len(years)-1
    for ri,y in enumerate(years,D0):
        tot_a=sum(wdf_a[wdf_a["Year"]==y][t].sum() for t in GROUP_A["types"] if t in wdf_a.columns)
        tot_b=sum(wdf_b[wdf_b["Year"]==y][t].sum() for t in GROUP_B["types"] if t in wdf_b.columns)
        ws.cell(ri,1,y); core.fmt(ws.cell(ri,1))
        for ci,tot,cap in [(2,tot_a,GROUP_A["cap_total"]),(4,tot_b,GROUP_B["cap_total"])]:
            cell=ws.cell(ri,ci,round(tot,1)); core.fmt(cell,number_format="#,##0.0")
            util=tot/cap if cap>0 else 0
            cell.fill=PatternFill("solid",start_color="FF4444" if util>1 else "FFD700" if util>=0.8 else "C6EFCE")
            if util>1: cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
        ws.cell(ri,3,GROUP_A["cap_total"]); core.fmt(ws.cell(ri,3),number_format="#,##0")
        ws.cell(ri,5,GROUP_B["cap_total"]); core.fmt(ws.cell(ri,5),number_format="#,##0")
    chart=BarChart(); chart.type="col"; chart.grouping="clustered"
    chart.title="Annual Demand — Mechanical vs Coating Labs"
    chart.y_axis.title="Samples/Year"; chart.x_axis.title="Year"
    chart.style=10; chart.height=15; chart.width=28
    cats=Reference(ws,min_col=1,min_row=D0,max_row=DE)
    first=Reference(ws,min_col=2,max_col=2,min_row=HR,max_row=DE)
    chart.add_data(first,titles_from_data=True); chart.set_categories(cats)
    for ci in [4,3,5]:
        ref=Reference(ws,min_col=ci,max_col=ci,min_row=HR,max_row=DE)
        chart.add_data(ref,titles_from_data=True)
    ws.add_chart(chart,f"A{DE+4}")
    # Monthly util line
    UR=DE+22
    core.section_hdr(ws,UR,"Monthly Utilization % (last 2 years)",10); UR+=1
    ws.cell(UR,1,"Month").font=Font(bold=True,name="Arial",size=9)
    ws.cell(UR,1).fill=PatternFill("solid",start_color="D9E1F2")
    plot_years=list(years[-2:]) if len(years)>=2 else list(years)
    col_a={y:2+2*i for i,y in enumerate(plot_years)}
    col_b={y:3+2*i for i,y in enumerate(plot_years)}
    for y in plot_years:
        for col,label in [(col_a[y],f"Mech {y}"),(col_b[y],f"Coating {y}")]:
            h=ws.cell(UR,col,label); h.font=Font(bold=True,name="Arial",size=8)
            h.fill=PatternFill("solid",start_color="D9E1F2")
            h.alignment=Alignment(horizontal="center"); core.cw(ws,col,14)
    core.cw(ws,1,10)
    for mi,month in enumerate(months):
        row=UR+1+mi
        ws.cell(row,1,month).font=Font(bold=True,name="Arial",size=9)
        ws.cell(row,1).fill=PatternFill("solid",start_color="EBF3FB")
        for y in plot_years:
            for wdf,grp,col in [(wdf_a,GROUP_A,col_a[y]),(wdf_b,GROUP_B,col_b[y])]:
                yd=wdf[wdf["Year"]==y].copy(); yd["_m"]=yd["Week"].apply(core.week_to_month)
                tot=sum(yd[yd["_m"]==mi+1][t].mean() for t in grp["types"] if t in yd.columns)
                util=tot/(grp["cap_total"]/52) if grp["cap_total"]>0 else 0
                if math.isnan(util): util=0.0
                cell=ws.cell(row,col,round(util,4)); cell.number_format="0.0%"
                cell.font=Font(name="Arial",size=9); cell.alignment=Alignment(horizontal="center")
    chart2=LineChart()
    chart2.title="Monthly Utilization — Mechanical vs Coating Labs"
    chart2.y_axis.title="Utilization %"; chart2.x_axis.title="Month"
    chart2.y_axis.numFmt="0%"; chart2.y_axis.scaling.min=0
    chart2.style=10; chart2.height=14; chart2.width=28
    all_cols=list(col_a.values())+list(col_b.values())
    first_col=min(all_cols)
    first_ref=Reference(ws,min_col=first_col,max_col=first_col,min_row=UR,max_row=UR+12)
    chart2.add_data(first_ref,titles_from_data=True)
    cats2=Reference(ws,min_col=1,min_row=UR+1,max_row=UR+12); chart2.set_categories(cats2)
    for col in all_cols:
        if col==first_col: continue
        ref=Reference(ws,min_col=col,max_col=col,min_row=UR,max_row=UR+12)
        chart2.add_data(ref,titles_from_data=True)
    ws.add_chart(chart2,f"A{UR+15}")


def write_all_labs(wb, wdf_a, wdf_b, years, types_a, types_b):
    ws=wb.create_sheet("All_Labs_Summary")
    core.banner(ws,1,"All Labs — Combined Annual Demand Summary",cols=20)
    all_types=types_a+types_b
    all_wdfs={t:wdf_a for t in types_a}; all_wdfs.update({t:wdf_b for t in types_b})
    all_caps={**GROUP_A["caps"],**GROUP_B["caps"]}
    HR=3; D0=4
    hdrs=["Year"]+all_types+["Mech Total","Coating Total","Grand Total"]
    for ci,h in enumerate(hdrs,1):
        bg="1F3864" if ci==1 else GROUP_A["color"] if ci<=1+len(types_a) else GROUP_B["color"] if ci<=1+len(all_types) else "C00000"
        core.hdr(ws.cell(HR,ci,h),bg=bg,sz=9); core.cw(ws,ci,13)
    for ri,y in enumerate(years,D0):
        ws.cell(ri,1,y); core.fmt(ws.cell(ri,1))
        mech_tot=coat_tot=0
        for ci,t in enumerate(all_types,2):
            wdf=all_wdfs[t]; val=wdf[wdf["Year"]==y][t].sum() if t in wdf.columns else 0
            cell=ws.cell(ri,ci,round(val,1)); core.fmt(cell,number_format="#,##0.0")
            cap=all_caps.get(t,1); util=val/cap if cap>0 else 0
            cell.fill=PatternFill("solid",start_color="FF4444" if util>1 else "FFD700" if util>=0.8 else "C6EFCE")
            if util>1: cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
            if t in types_a: mech_tot+=val
            else: coat_tot+=val
        lc=1+len(all_types)
        mc=ws.cell(ri,lc+1,round(mech_tot,1)); core.fmt(mc,number_format="#,##0.0",bold=True); mc.fill=PatternFill("solid",start_color="BDD7EE")
        cc=ws.cell(ri,lc+2,round(coat_tot,1)); core.fmt(cc,number_format="#,##0.0",bold=True); cc.fill=PatternFill("solid",start_color="C6EFCE")
        gc=ws.cell(ri,lc+3,round(mech_tot+coat_tot,1)); core.fmt(gc,number_format="#,##0.0",bold=True); gc.fill=PatternFill("solid",start_color="D9D9D9")
    chart=BarChart(); chart.type="col"; chart.grouping="clustered"
    chart.title="All Labs — Annual Demand"; chart.y_axis.title="Samples/Year"
    chart.style=10; chart.height=16; chart.width=34
    cats=Reference(ws,min_col=1,min_row=D0,max_row=D0+len(years)-1)
    first=Reference(ws,min_col=2,max_col=2,min_row=HR,max_row=D0+len(years)-1)
    chart.add_data(first,titles_from_data=True); chart.set_categories(cats)
    for ci in range(3,2+len(all_types)):
        ref=Reference(ws,min_col=ci,max_col=ci,min_row=HR,max_row=D0+len(years)-1)
        chart.add_data(ref,titles_from_data=True)
    ws.add_chart(chart,f"A{D0+len(years)+4}")


# ────────────────────────────────────────────────────────────
#  MAIN GENERATOR
# ────────────────────────────────────────────────────────────

def generate(file_paths, caps_a, caps_b, theme, progress_cb=None):
    def p(msg):
        if progress_cb: progress_cb(msg)

    p(f"Loading {len(file_paths)} file(s)…")
    merged_df, col_map, errors, file_log = merge_files(file_paths)
    if errors:
        raise ValueError("\n".join(errors))

    tc = col_map["type"]
    df_a = merged_df[merged_df[tc].isin(GROUP_A["types"])].copy()
    df_b = merged_df[merged_df[tc].isin(GROUP_B["types"])].copy()
    if df_a.empty:
        raise ValueError(f"No data for Mechanical labs {GROUP_A['types']}.")
    if df_b.empty:
        raise ValueError(f"No data for Coating labs {GROUP_B['types']}.")

    p("Computing weekly data…")
    wdf_a,udf_a,types_a,years_a = core.build_weekly(df_a, col_map, caps_a)
    wdf_b,udf_b,types_b,years_b = core.build_weekly(df_b, col_map, caps_b)
    years = sorted(set(list(years_a)) | set(list(years_b)))

    p("Building workbook…")
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    p("Sheet 1/8 — Data Sources…")
    write_data_sources(wb, file_paths, file_log)
    p("Sheet 2/8 — Overview…")
    write_overview(wb, wdf_a, wdf_b, years)
    p("Sheet 3/8 — All Labs Summary…")
    write_all_labs(wb, wdf_a, wdf_b, years, types_a, types_b)
    p("Sheet 4/8 — Comparison Charts…")
    write_comparison_chart(wb, wdf_a, wdf_b, years)
    p("Sheet 5/8 — Mech Utilization…")
    core.write_utilization_sheet(wb, udf_a, types_a, theme)
    wb.worksheets[-1].title = "Mech_Utilization"
    p("Sheet 6/8 — Coating Utilization…")
    core.write_utilization_sheet(wb, udf_b, types_b, theme)
    wb.worksheets[-1].title = "Coating_Utilization"
    p("Sheet 7/8 — Mech YoY…")
    core.write_yoy_chart(wb, wdf_a, types_a, years_a)
    wb.worksheets[-1].title = "Mech_YoY"
    p("Sheet 8/8 — Coating YoY…")
    core.write_yoy_chart(wb, wdf_b, types_b, years_b)
    wb.worksheets[-1].title = "Coating_YoY"

    p("Saving…")
    out = core.save_workbook(wb, file_paths[0], OUTPUT_FILE)
    return out, file_log


# ────────────────────────────────────────────────────────────
#  GUI
# ────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"Lab Comparison Dashboard")
        self.geometry("800x760")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self._file_list  = []
        self.caps_a      = {k: tk.IntVar(value=v) for k,v in GROUP_A["caps"].items()}
        self.caps_b      = {k: tk.IntVar(value=v) for k,v in GROUP_B["caps"].items()}
        self.theme_var   = tk.StringVar(value=list(core.COLOR_THEMES.keys())[0])
        self.progress_var= tk.StringVar(value="")
        self._build()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build(self):
        hdr=tk.Frame(self,bg="#3D1F6E",height=74); hdr.pack(fill="x")
        tk.Label(hdr,text="Lab Comparison Dashboard",bg="#3D1F6E",fg="white",
                 font=("Arial",14,"bold")).pack(pady=8)
        tk.Label(hdr,text="Mechanical (LCF + Creep)  vs  Coating (Cold Spray + HVOF + Plasma)",
                 bg="#3D1F6E",fg="#D9B3FF",font=("Arial",9)).pack()

        main=tk.Frame(self,bg="#F0F4F8",padx=20,pady=8); main.pack(fill="both",expand=True)

        self._sec(main,"Step 1 — Upload Files  (one or more Excel files, data merged automatically)")
        fcard=tk.Frame(main,bg="#FFFFFF",bd=1,relief="groove",padx=10,pady=8)
        fcard.pack(fill="x",pady=(0,4))
        btn_row=tk.Frame(fcard,bg="#FFFFFF"); btn_row.pack(fill="x")
        self._btn(btn_row,"+ Add File(s)",self._add_files,color="#3D1F6E").pack(side="left")
        self._btn(btn_row,"Clear All",self._clear_files,color="#888888").pack(side="left",padx=8)
        tk.Label(btn_row,text="Tip: Works with 1 combined file or 2+ separate files",
                 bg="#FFFFFF",fg="#888",font=("Arial",8)).pack(side="left",padx=6)
        lf=tk.Frame(fcard,bg="#FFFFFF"); lf.pack(fill="x",pady=(6,0))
        self._lb=tk.Listbox(lf,height=4,font=("Arial",9),bg="#F8F8F8",relief="flat",bd=1)
        self._lb.pack(side="left",fill="x",expand=True)
        sb=ttk.Scrollbar(lf,orient="vertical",command=self._lb.yview)
        sb.pack(side="right",fill="y"); self._lb.config(yscrollcommand=sb.set)
        self._btn(lf,"Remove\nSelected",self._remove_sel,color="#C00000",size=8).pack(side="right",padx=(4,0))
        self._cnt_lbl=tk.Label(fcard,text="No files added yet.",bg="#FFFFFF",fg="#888",
                                font=("Arial",9,"italic")); self._cnt_lbl.pack(anchor="w",pady=(4,0))

        ttk.Separator(main).pack(fill="x",pady=6)
        self._sec(main,"Mechanical Lab Capacities  (samples/year)")
        af=tk.Frame(main,bg="#F0F4F8"); af.pack(fill="x",pady=(0,6))
        for i,(name,var) in enumerate(self.caps_a.items()):
            card=tk.Frame(af,bg="#EBF3FB",bd=1,relief="groove",padx=12,pady=8)
            card.grid(row=0,column=i,padx=8,sticky="ew"); af.columnconfigure(i,weight=1)
            tk.Label(card,text=name,bg="#EBF3FB",font=("Arial",10,"bold"),fg="#1A4E8A").pack(anchor="w")
            tk.Spinbox(card,from_=1,to=9999,textvariable=var,width=10,
                       font=("Arial",11),relief="flat").pack(anchor="w",pady=(3,0))

        ttk.Separator(main).pack(fill="x",pady=6)
        self._sec(main,"Coating Lab Capacities  (samples/year — combined = 350)")
        bf=tk.Frame(main,bg="#F0F4F8"); bf.pack(fill="x",pady=(0,6))
        for i,(name,var) in enumerate(self.caps_b.items()):
            card=tk.Frame(bf,bg="#EBF9EC",bd=1,relief="groove",padx=12,pady=8)
            card.grid(row=0,column=i,padx=6,sticky="ew"); bf.columnconfigure(i,weight=1)
            tk.Label(card,text=name,bg="#EBF9EC",font=("Arial",10,"bold"),fg="#1F5C1A").pack(anchor="w")
            tk.Spinbox(card,from_=1,to=9999,textvariable=var,width=10,
                       font=("Arial",11),relief="flat").pack(anchor="w",pady=(3,0))

        ttk.Separator(main).pack(fill="x",pady=6)
        self._sec(main,"Color Theme")
        tf=tk.Frame(main,bg="#F0F4F8"); tf.pack(fill="x",pady=(0,6))
        for i,name in enumerate(core.COLOR_THEMES):
            tk.Radiobutton(tf,text=name,variable=self.theme_var,value=name,
                           bg="#F0F4F8",font=("Arial",9)).grid(row=0,column=i,padx=8,sticky="w")

        ttk.Separator(main).pack(fill="x",pady=6)
        self._pl=tk.Label(main,textvariable=self.progress_var,bg="#F0F4F8",
                           fg="#3D1F6E",font=("Arial",9,"italic")); self._pl.pack(anchor="w")
        self._pb=ttk.Progressbar(main,mode="indeterminate",length=740)
        self._pb.pack(fill="x",pady=4)
        self.gen_btn=self._btn(main,"Generate Comparison Dashboard",
                               self._run,color="#3D1F6E",size=11,pady=13)
        self.gen_btn.pack(pady=8,fill="x")
        ftr=tk.Frame(self,bg="#DDE5EF",height=26); ftr.pack(fill="x",side="bottom")
        tk.Label(ftr,text=f"Output -> Lab_Dashboard_Output/{OUTPUT_FILE}",
                 bg="#DDE5EF",fg="#666",font=("Arial",8)).pack(pady=5)

    def _sec(self,p,title):
        tk.Label(p,text=title,bg="#F0F4F8",font=("Arial",10,"bold"),
                 fg="#3D1F6E").pack(anchor="w",pady=(6,2))

    def _btn(self,p,text,cmd,color="#3D1F6E",size=9,pady=8):
        return tk.Button(p,text=text,command=cmd,bg=color,fg="white",
                         activebackground=color,font=("Arial",size,"bold"),
                         relief="flat",cursor="hand2",padx=14,pady=pady)

    def _add_files(self):
        paths=filedialog.askopenfilenames(
            title="Select Lab Data Excel File(s)",
            filetypes=[("Excel files","*.xlsx *.xls"),("All files","*.*")])
        added=0
        for path in paths:
            if path not in self._file_list:
                self._file_list.append(path)
                self._lb.insert(tk.END,f"  {os.path.basename(path)}")
                added+=1
        if added:
            self._upd_cnt(); self.progress_var.set("")

    def _remove_sel(self):
        sel=self._lb.curselection()
        if not sel: messagebox.showinfo("Select First","Click a file first."); return
        idx=sel[0]; self._file_list.pop(idx); self._lb.delete(idx); self._upd_cnt()

    def _clear_files(self):
        self._file_list.clear(); self._lb.delete(0,tk.END)
        self._upd_cnt(); self.progress_var.set("")

    def _upd_cnt(self):
        n=len(self._file_list)
        if n==0: self._cnt_lbl.config(text="No files added yet.",fg="#888")
        elif n==1: self._cnt_lbl.config(text="1 file loaded.",fg="#1F5C1A")
        else: self._cnt_lbl.config(text=f"{n} files loaded — data will be merged automatically.",fg="#1F5C1A")

    def _run(self):
        if not self._file_list:
            messagebox.showerror("No Files","Please add at least one Excel file first."); return
        caps_a={k:v.get() for k,v in self.caps_a.items()}
        caps_b={k:v.get() for k,v in self.caps_b.items()}
        theme=core.COLOR_THEMES[self.theme_var.get()]
        self.gen_btn.config(state="disabled",text="Generating...")
        self._pb.start(10)
        def task():
            try:
                out,log=generate(list(self._file_list),caps_a,caps_b,theme,
                    progress_cb=lambda m:self.after(0,lambda msg=m:self.progress_var.set(msg)))
                self.after(0,lambda:self._ok(out,log))
            except Exception as e:
                self.after(0,lambda err=str(e):self._err(err))
        threading.Thread(target=task,daemon=True).start()

    def _ok(self,out,log):
        self._pb.stop(); self.gen_btn.config(state="normal",text="Generate Comparison Dashboard")
        self.progress_var.set("Done!")
        msg=f"Dashboard created!\n\n{out}\n\n8 sheets generated.\n\nFile load log:\n"+"\n".join(log)
        messagebox.showinfo("Success",msg)

    def _err(self,e):
        self._pb.stop(); self.gen_btn.config(state="normal",text="Generate Comparison Dashboard")
        self.progress_var.set("Error"); messagebox.showerror("Error",e)


if __name__ == "__main__":
    App().mainloop()
