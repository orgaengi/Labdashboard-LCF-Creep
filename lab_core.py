"""
lab_core.py  — Shared utilities for all 3 Lab Dashboard tools
"""

import os
import math
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
#  SEASON WEIGHTS  (52-week, sum=52)
# ─────────────────────────────────────────────
def _make_seasonal():
    w = []
    for i in range(1, 53):
        v = (1.0
             + 0.22 * math.sin(2 * math.pi * (i - 8)  / 52)
             + 0.12 * math.sin(4 * math.pi * (i - 4)  / 52)
             + 0.06 * math.sin(6 * math.pi * (i - 2)  / 52))
        w.append(max(0.55, v))
    s = sum(w)
    return [x * 52 / s for x in w]

SEASONAL = _make_seasonal()

COLUMN_ALIASES = {
    "year":  ["year", "yr", "fiscal year", "fy"],
    "type":  ["type", "process", "category", "lab", "process type", "lab type"],
    "value": ["value", "count", "samples", "qty", "quantity", "volume",
              "total", "no of samples", "number"],
}

COLOR_THEMES = {
    "Default (Red/Yellow/Green)": {
        "high": "FF4444", "medium": "FFD700", "low": "70AD47",
        "high_font": "FFFFFF", "medium_font": "333333", "low_font": "FFFFFF",
    },
    "Blue Gradient": {
        "high": "1F4E79", "medium": "2E75B6", "low": "BDD7EE",
        "high_font": "FFFFFF", "medium_font": "FFFFFF", "low_font": "333333",
    },
    "Pastel": {
        "high": "FF9999", "medium": "FFEB99", "low": "C6EFCE",
        "high_font": "333333", "medium_font": "333333", "low_font": "333333",
    },
    "Teal/Coral": {
        "high": "E05050", "medium": "E8A838", "low": "2AADA8",
        "high_font": "FFFFFF", "medium_font": "333333", "low_font": "FFFFFF",
    },
    "Purple Haze": {
        "high": "7030A0", "medium": "C55A11", "low": "4BACC6",
        "high_font": "FFFFFF", "medium_font": "FFFFFF", "low_font": "FFFFFF",
    },
    "Monochrome": {
        "high": "262626", "medium": "737373", "low": "BFBFBF",
        "high_font": "FFFFFF", "medium_font": "FFFFFF", "low_font": "333333",
    },
}

CURRENT_YEAR = datetime.date.today().year
CURRENT_WEEK = datetime.date.today().isocalendar()[1]

# ─────────────────────────────────────────────
#  STYLE HELPERS
# ─────────────────────────────────────────────
def bdr(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg="1F3864", fg="FFFFFF", sz=10, bold=True):
    cell.font      = Font(bold=bold, color=fg, name="Arial", size=sz)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = bdr()

def fmt(cell, number_format=None, bold=False, center=True, sz=9, color="000000"):
    cell.font   = Font(name="Arial", size=sz, bold=bold, color=color)
    cell.border = bdr()
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        cell.number_format = number_format

def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def banner(ws, row, text, bg="1F3864", fg="FFFFFF", sz=13, height=36, cols=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font      = Font(bold=True, size=sz, name="Arial", color=fg)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height

def section_hdr(ws, row, text, col_end, bg="2E75B6"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row, 1, text)
    c.font      = Font(bold=True, size=10, name="Arial", color="FFFFFF")
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

# ─────────────────────────────────────────────
#  DATA LAYER
# ─────────────────────────────────────────────
def detect_columns(df):
    mapping = {}
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for cl, co in cols_lower.items():
                if alias in cl:
                    mapping[field] = co
                    break
            if field in mapping:
                break
    return mapping

def validate_data(df, col_map, allowed_types=None):
    """Returns (errors, warnings)."""
    errors, warns = [], []
    missing = [f for f in ["year", "type", "value"] if f not in col_map]
    if missing:
        errors.append(
            f"Required column(s) not found: {', '.join(missing).upper()}\n"
            f"Detected: {list(df.columns)}"
        )
        return errors, warns

    yc, tc, vc = col_map["year"], col_map["type"], col_map["value"]
    null_n = df[[yc, tc, vc]].isnull().sum().sum()
    if null_n:
        warns.append(f"{null_n} null value(s) found — those rows will be skipped.")

    df2 = df.dropna(subset=[yc, tc, vc])
    if df2.empty:
        errors.append("No valid rows after removing nulls.")
        return errors, warns

    nums = pd.to_numeric(df2[vc], errors="coerce")
    if (nums < 0).sum():
        errors.append(f"Negative values found in '{vc}'. All values must be ≥ 0.")
    if (nums.isna()).sum():
        warns.append("Some non-numeric values in value column — treated as 0.")

    if allowed_types:
        found = set(df2[tc].str.strip().str.lower().unique())
        expected = {t.lower() for t in allowed_types}
        missing_t = expected - found
        if missing_t:
            warns.append(
                f"These lab types were not found in the file: {missing_t}. "
                "They will be treated as 0."
            )
    return errors, warns

def _is_wide_format(df, allowed_types):
    """Check if df is wide format: has Year + lab-type columns directly."""
    cols_lower = {c.lower().strip(): c for c in df.columns}
    allowed_lower = [t.lower() for t in allowed_types]
    year_found = any(a in cl for cl in cols_lower for a in COLUMN_ALIASES["year"])
    type_cols  = [c for al in allowed_lower for cl, c in cols_lower.items() if al == cl]
    return year_found and len(type_cols) >= 1 and "type" not in cols_lower and "value" not in cols_lower


def _wide_to_long(df, allowed_types):
    """Melt wide-format (Year, Week?, Lab1, Lab2..) to long format (Year, Week, Type, Value)."""
    cols_lower = {c.lower().strip(): c for c in df.columns}
    # Identify year col
    year_col = None
    for alias in COLUMN_ALIASES["year"]:
        for cl, co in cols_lower.items():
            if alias in cl:
                year_col = co; break
        if year_col: break
    # Identify week col
    week_col = None
    for cl, co in cols_lower.items():
        if "week" in cl or "wk" in cl:
            week_col = co; break
    # Identify lab type columns
    allowed_lower_map = {t.lower(): t for t in allowed_types}
    type_cols = []
    for cl, co in cols_lower.items():
        if cl in allowed_lower_map:
            type_cols.append((co, allowed_lower_map[cl]))
    rows = []
    for _, row in df.iterrows():
        year = row[year_col]
        # Skip summary / total rows
        try:
            year = int(float(year))
        except (ValueError, TypeError):
            continue
        week = row[week_col] if week_col else None
        try:
            week = int(float(week)) if week is not None and not pd.isna(week) else None
        except (ValueError, TypeError):
            week = None
        for orig_col, canonical_type in type_cols:
            val = row[orig_col]
            try:
                val = float(val)
            except:
                val = 0.0
            if pd.isna(val): val = 0.0
            r = {"Year": year, "Type": canonical_type, "Value": val}
            if week is not None:
                r["Week"] = week
            rows.append(r)
    return pd.DataFrame(rows)


def _scan_header_row(df_raw):
    """
    Given a DataFrame read with header=None, find the row index that looks
    most like a real header (contains known keyword aliases).
    Returns (header_row_idx, df_with_correct_header).
    """
    all_aliases = set()
    for aliases in COLUMN_ALIASES.values():
        all_aliases.update(aliases)
    all_aliases.update(["week", "wk"])

    best_row, best_score = 0, 0
    for i, row in df_raw.iterrows():
        score = sum(
            1 for cell in row
            if isinstance(cell, str) and
               any(a in cell.lower().strip() for a in all_aliases)
        )
        if score > best_score:
            best_score, best_row = score, i

    if best_score == 0:
        # No keyword row found — assume row 0 is header
        best_row = 0

    # Rebuild df with that row as header
    new_df = df_raw.iloc[best_row + 1:].copy()
    new_df.columns = [str(c).strip() if not pd.isna(c) else f"_col{i}"
                      for i, c in enumerate(df_raw.iloc[best_row])]
    new_df = new_df.reset_index(drop=True)
    return new_df


def _try_all_sheets(path):
    """
    Try to read each sheet in the workbook. Return the first sheet whose
    DataFrame (after header scan) contains recognisable column keywords.
    Falls back to first sheet if none match.
    """
    all_aliases = set()
    for aliases in COLUMN_ALIASES.values():
        all_aliases.update(aliases)

    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return None, "Cannot open file"

    sheet_names = xl.sheet_names

    # Priority: if "Weekly_Planner" exists, try it first
    if "Weekly_Planner" in sheet_names:
        sheet_names = ["Weekly_Planner"] + [s for s in sheet_names if s != "Weekly_Planner"]

    best_df, best_score = None, -1
    for sheet in sheet_names:
        try:
            raw = pd.read_excel(xl, sheet_name=sheet, header=None)
        except Exception:
            continue
        if raw.empty or len(raw) < 2:
            continue
        # Score: how many cells in first 10 rows match known aliases
        score = 0
        for _, row in raw.head(10).iterrows():
            for cell in row:
                if isinstance(cell, str):
                    if any(a in cell.lower().strip() for a in all_aliases):
                        score += 1
        if score > best_score:
            best_score, best_df = score, raw

    if best_df is None:
        return None, "No readable sheet found"
    return best_df, None


def _fuzzy_match_type(cell_value, allowed_types):
    """
    Return the canonical allowed_type that best matches cell_value, or None.
    Strategy (in order):
      1. Exact match (case-insensitive, stripped)
      2. All words of cell_value appear in allowed_type (or vice versa)
      3. Any word of cell_value ≥4 chars appears in allowed_type
    """
    if not isinstance(cell_value, str):
        return None
    v = cell_value.lower().strip()
    for t in allowed_types:
        if v == t.lower():
            return t
    for t in allowed_types:
        tl = t.lower()
        v_words = set(v.split())
        t_words = set(tl.split())
        if v_words and v_words <= t_words:          # all cell words in type
            return t
        if t_words and t_words <= v_words:          # all type words in cell
            return t
    for t in allowed_types:
        tl = t.lower()
        for word in v.split():
            if len(word) >= 4 and word in tl:
                return t
    return None


def _is_monthly_block_format(path):
    """
    Detect the monthly-block format:
      - Sheet names are years (e.g. '2024', '2025')
      - No column headers; repeating 5-row monthly blocks:
          [date_row]
          TOTAL RIGS RUNNING   | value
          TOTAL SLOTS AVAILABLE| value
          TOTAL SAMPLES REMOVED| value
          [blank row]
    Returns (True/False, ExcelFile object, list_of_year_sheets).
    """
    import re as _re
    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return False, None, []

    # Sheet names are expected to represent a year, but real-world workbooks
    # often label them "YEAR 2024", "FY 2024", "2024 Data", etc. rather than
    # a bare "2024". Search for a 20xx year ANYWHERE in the name instead of
    # requiring the whole name to be exactly 4 digits — this was the cause
    # of "No valid rows after removing nulls" for files using a "YEAR "
    # prefix even though the underlying block data was perfectly valid.
    YEAR_RE = _re.compile(r'(20\d{2})')
    year_sheets = [s for s in xl.sheet_names if YEAR_RE.search(str(s).strip())]
    if not year_sheets:
        return False, xl, []

    # Verify at least one sheet has the SAMPLES REMOVED keyword
    for sheet in year_sheets[:2]:
        try:
            raw = pd.read_excel(xl, sheet_name=sheet, header=None)
            col0 = raw.iloc[:, 0].astype(str).str.upper()
            if col0.str.contains("SAMPLES REMOVED", na=False).any():
                return True, xl, year_sheets
        except Exception:
            continue
    return False, xl, []


def _infer_type_from_sheet(sheet_name, allowed_types):
    """
    Extract a lab type from a sheet name that contains both a type label
    and a year, e.g. "LCF 2024", "Cold Spray 2025", "HVOF 2024".

    Strategy (in order):
      1. Remove the year (20xx) and filler words, then exact-match the remainder
      2. All words of a type appear in the remainder
      3. Any word of a type (any length) is a substring of the remainder
    Returns the matched canonical type string, or None.
    """
    import re as _re
    year_re = _re.compile(r'20\d{2}')
    remainder = year_re.sub('', str(sheet_name)).strip().lower()
    # strip common filler words
    for filler in ('data', 'lab', 'report', 'monthly', 'fy', 'year', 'sheet'):
        remainder = remainder.replace(filler, '')
    remainder = remainder.strip(' _-')

    if not remainder:
        return None   # sheet name was purely a year (e.g. "2024") → no type

    # 1. Exact full-name match
    for t in allowed_types:
        if t.lower().strip() == remainder:
            return t

    # 2. All words of type present in remainder
    for t in allowed_types:
        t_words = t.lower().split()
        if t_words and all(w in remainder for w in t_words):
            return t

    # 3. Any single word of the type is a substring (no length floor — covers LCF, CS, etc.)
    for t in allowed_types:
        for w in t.lower().split():
            if w and w in remainder:
                return t

    return None


def _parse_monthly_blocks(xl, year_sheets, allowed_types, source_path=None):
    """
    Parse the monthly-block format.

    Sheet naming conventions:
      Single-type tool: plain year sheets "2024", "FY 2024", "YEAR 2024"
      Multi-type tool:  type + year sheets "LCF 2024", "Cold Spray 2025"

    Each ~5-row block per month:
        row 0: date  |  row 1: RIGS RUNNING  |  row 2: SLOTS AVAILABLE
        row 3: SAMPLES REMOVED  |  row 4: blank

    Returns (annual_df, warns) — annual_df columns: Year | Type | Value
    """
    warns = []; monthly_rows = []
    import re as _re
    _YEAR_EXTRACT = _re.compile(r'(20\d{2})')

    for sheet in year_sheets:
        m = _YEAR_EXTRACT.search(str(sheet).strip())
        if not m:
            warns.append(f"Sheet '{sheet}' skipped — no valid year found in name.")
            continue
        year = int(m.group(1))
        sheet_type = _infer_type_from_sheet(sheet, allowed_types)

        try:
            raw = pd.read_excel(xl, sheet_name=sheet, header=None)
        except Exception as e:
            warns.append(f"Could not read sheet '{sheet}': {e}")
            continue

        i = 0
        while i < len(raw):
            row   = raw.iloc[i]
            cell0 = row.iloc[0]
            month_num = None
            if hasattr(cell0, 'month'):
                month_num = cell0.month
            elif isinstance(cell0, str) and cell0.strip():
                dt = pd.to_datetime(cell0, errors='coerce')
                if not pd.isna(dt):
                    month_num = dt.month

            if month_num is not None:
                block = {'Year': year, 'Month': month_num,
                         'Rigs': None, 'Slots': None, 'Removed': None,
                         'SheetType': sheet_type}
                j = i + 1
                while j < len(raw) and j <= i + 4:
                    r2  = raw.iloc[j]
                    lbl = str(r2.iloc[0]).upper().strip() if not pd.isna(r2.iloc[0]) else ""
                    val = r2.iloc[1] if len(r2) > 1 else None
                    try:
                        val = float(val) if val is not None and not pd.isna(val) else None
                    except (ValueError, TypeError):
                        val = None
                    if 'SAMPLES REMOVED' in lbl and val is not None:
                        block['Removed'] = val
                    elif 'SLOTS AVAILABLE' in lbl and val is not None:
                        block['Slots'] = val
                    elif 'RIGS RUNNING' in lbl and val is not None:
                        block['Rigs'] = val
                    j += 1
                if block['Removed'] is not None:
                    monthly_rows.append(block)
                i = j
            else:
                i += 1

    if not monthly_rows:
        return pd.DataFrame(), ["No monthly data blocks found in the file."]

    detail = pd.DataFrame(monthly_rows)

    # ── Determine type assignment ─────────────────────────────────────────────
    has_sheet_types = (
        'SheetType' in detail.columns and
        detail['SheetType'].notna().any() and
        len(allowed_types) > 1
    )

    if has_sheet_types:
        # Multi-type: each sheet carries a type label (e.g. "LCF 2024")
        unmatched = detail['SheetType'].isna().sum()
        if unmatched:
            detail['SheetType'] = detail['SheetType'].fillna(allowed_types[0])
            warns.append(
                f"{unmatched} block(s) had no recognisable type in the sheet name "
                f"— assigned to '{allowed_types[0]}'. "
                "Rename sheets to include the lab type (e.g. 'LCF 2024', 'Creep 2024')."
            )
        annual = (
            detail.groupby(['Year', 'SheetType'])['Removed']
            .sum().reset_index()
            .rename(columns={'Removed': 'Value', 'SheetType': 'Type'})
        )

    elif len(allowed_types) == 1:
        # Single-type tool — original OHC path
        annual = (detail.groupby('Year')['Removed']
                  .sum().reset_index().rename(columns={'Removed': 'Value'}))
        annual['Type'] = allowed_types[0]

    else:
        # Multiple types but no sheet labels — fall back to filename inference
        type_name = None
        if source_path:
            fname_lower = os.path.basename(source_path).lower()
            fname_clean = fname_lower.replace('_', ' ').replace('-', ' ')
            for t in allowed_types:
                if all(w in fname_clean for w in t.lower().split()):
                    type_name = t; break
            if type_name is None:
                for t in allowed_types:
                    for w in t.lower().split():
                        if len(w) >= 4 and w in fname_clean:
                            type_name = t; break
                    if type_name: break
        if type_name:
            warns.append(f"Inferred type '{type_name}' from filename.")
        else:
            type_name = allowed_types[0]
            warns.append(
                f"Could not infer lab type. All data assigned to '{type_name}'. "
                "Add the type name to each sheet (e.g. 'LCF 2024', 'Creep 2024')."
            )
        annual = (detail.groupby('Year')['Removed']
                  .sum().reset_index().rename(columns={'Removed': 'Value'}))
        annual['Type'] = type_name

    # ── Partial-year warnings ─────────────────────────────────────────────────
    partial = [f"{yr} ({len(g)} months)" for yr, g in detail.groupby('Year') if len(g) < 12]
    if partial:
        warns.append(f"Partial year(s) detected: {', '.join(partial)}.")

    annual['Year']  = annual['Year'].astype(int)
    annual['Value'] = annual['Value'].fillna(0)
    return annual[['Year', 'Type', 'Value']], warns




def load_and_filter(path, allowed_types):
    """
    Load Excel (any format, any sheet, any header row), return
    (df, col_map, errors, warns).

    Handles:
    - monthly-block format (year-named sheets, keyword/value pairs)
    - Junk / title rows above the real header
    - Data on any sheet (scans all, picks best)
    - Wide format (Year | LabType1 | LabType2 ...) and long format
    - Partial / fuzzy type name matching
    - Mixed-case type names and extra whitespace
    - Blank rows, subtotal rows, non-numeric values
    - Year values like 'FY2024', '2024-25', dates
    - Extra/junk columns (Notes, Remarks, etc.)
    """
    # ── 0. monthly-block format (highest priority) ─
    is_block, xl_obj, year_sheets = _is_monthly_block_format(path)
    if is_block:
        annual_df, block_warns = _parse_monthly_blocks(xl_obj, year_sheets, allowed_types, source_path=path)
        if annual_df is not None and not annual_df.empty:
            col_map = {'year': 'Year', 'type': 'Type', 'value': 'Value'}
            return annual_df, col_map, [], block_warns
        # If parse failed, fall through to general parser with warnings
        block_warns.insert(0, "Monthly-block format detected but parsing returned no data — trying general parser.")
    else:
        block_warns = []

    # ── 1. Find the best sheet ────────────────────────────
    raw_df, sheet_err = _try_all_sheets(path)
    if raw_df is None:
        return None, {}, [f"Cannot open file: {sheet_err}"], block_warns

    # ── 2. Find the real header row within the sheet ──────
    df = _scan_header_row(raw_df)

    if df.empty:
        return None, {}, ["File is empty or has no data rows."], []

    # ── 3. Detect wide vs long, convert if needed ─────────
    if _is_wide_format(df, allowed_types):
        df = _wide_to_long(df, allowed_types)
    else:
        # Wide format with fuzzy column names?
        # Check if any column fuzzy-matches a lab type
        fuzzy_wide_cols = []
        cols_lower = {c.lower().strip(): c for c in df.columns}
        for cl, co in cols_lower.items():
            matched = _fuzzy_match_type(cl, allowed_types)
            if matched:
                fuzzy_wide_cols.append((co, matched))
        if fuzzy_wide_cols:
            # Rename those columns to canonical names and treat as wide
            rename_map = {orig: canon for orig, canon in fuzzy_wide_cols}
            df = df.rename(columns=rename_map)
            df = _wide_to_long(df, allowed_types)

    # ── 4. Detect standard columns ────────────────────────
    col_map = detect_columns(df)

    # If Type column found, apply fuzzy matching on its values
    if "type" in col_map:
        tc = col_map["type"]
        df[tc] = df[tc].astype(str).str.strip()
        def _map_type(v):
            if not isinstance(v, str) or not v.strip():
                return v
            exact_lower = {t.lower(): t for t in allowed_types}
            vl = v.lower().strip()
            if vl in exact_lower:
                return exact_lower[vl]
            return _fuzzy_match_type(v, allowed_types) or v
        df[tc] = df[tc].apply(_map_type)

    # ── 5. Validate ───────────────────────────────────────
    errors, warns = validate_data(df, col_map, allowed_types)
    if errors:
        return df, col_map, errors, warns

    # ── 6. Final filter to allowed types ─────────────────
    tc = col_map["type"]
    df[tc] = df[tc].astype(str).str.strip()
    allowed_lower = {t.lower(): t for t in allowed_types}
    df["_type_lower"] = df[tc].str.lower()
    df = df[df["_type_lower"].isin(allowed_lower)]
    df[tc] = df["_type_lower"].map(allowed_lower)
    df = df.drop(columns=["_type_lower"])

    if df.empty:
        return df, col_map, [
            f"No rows matched the expected lab types: {allowed_types}. "
            f"Types found in file: {list(df[tc].unique()) if tc in df.columns else 'none'}"
        ], warns

    return df, col_map, [], warns

def build_weekly(df, col_map, capacities):
    """
    Annual → weekly with seasonal pattern.
    Returns weekly_df, util_df, types, years.
    """
    yc, tc, vc = col_map["year"], col_map["type"], col_map["value"]
    df = df.dropna(subset=[yc, tc, vc]).copy()
    df[vc]  = pd.to_numeric(df[vc], errors="coerce").fillna(0)
    df[yc]  = df[yc].astype(int)

    years = sorted(df[yc].unique())
    types = sorted(df[tc].unique())

    rows_w, rows_u = [], []
    for year in years:
        yd     = df[df[yc] == year]
        annual = {t: yd[yd[tc] == t][vc].sum() for t in types}
        for wi, week in enumerate(range(1, 53)):
            rw = {"Year": year, "Week": week}
            ru = {"Year": year, "Week": week}
            for t in types:
                dem     = annual[t] * SEASONAL[wi] / 52
                cap     = capacities.get(t, 1)
                cap_wk  = cap / 52          # ← weekly capacity (annual ÷ 52)
                rw[t]   = round(dem, 3)
                ru[t]   = round(dem / cap_wk if cap_wk > 0 else 0, 4)  # util vs weekly cap
            rows_w.append(rw)
            rows_u.append(ru)

    return pd.DataFrame(rows_w), pd.DataFrame(rows_u), types, years

def week_to_month(w):
    return min(12, math.ceil(w * 12 / 52))

# ─────────────────────────────────────────────
#  SHARED SHEET WRITERS
# ─────────────────────────────────────────────

def write_weekly_planner(wb, weekly_df, types, title_text):
    ws      = wb.create_sheet("Weekly_Planner")
    headers = ["Year", "Week"] + types
    banner(ws, 1, title_text, cols=len(headers))
    for c, h in enumerate(headers, 1):
        hdr(ws.cell(3, c, h))
        cw(ws, c, 15 if c == 1 else 13)

    years = sorted(weekly_df["Year"].unique())
    fills = [PatternFill("solid", start_color="EBF3FB"),
             PatternFill("solid", start_color="FFFFFF")]
    parity = {int(y): i % 2 for i, y in enumerate(years)}

    for r, row in enumerate(weekly_df.itertuples(index=False), 4):
        vals = list(row)
        yr   = int(vals[0])
        fill = fills[parity.get(yr, 0)]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r, c, val)
            fmt(cell, number_format="0.000" if c > 2 else None)
            cell.fill = fill

    last = weekly_df.shape[0] + 4
    ws.cell(last, 1, "TOTAL").font = Font(bold=True, name="Arial", size=9,
                                          color="FFFFFF")
    ws.cell(last, 1).fill = PatternFill("solid", start_color="1F3864")
    ws.cell(last, 2).fill = PatternFill("solid", start_color="1F3864")
    for c, t in enumerate(types, 3):
        cl   = get_column_letter(c)
        cell = ws.cell(last, c, f"=SUM({cl}4:{cl}{last-1})")
        fmt(cell, number_format="#,##0.000", bold=True)
        cell.fill = PatternFill("solid", start_color="D9E1F2")

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"


def write_utilization_sheet(wb, util_df, types, theme):
    ws      = wb.create_sheet("Utilization")
    headers = ["Year", "Week"] + types
    banner(ws, 1, "Utilization (Color-Coded)", cols=len(headers))

    hf = PatternFill("solid", start_color=theme["high"])
    mf = PatternFill("solid", start_color=theme["medium"])
    lf = PatternFill("solid", start_color=theme["low"])

    for c, h in enumerate(headers, 1):
        hdr(ws.cell(3, c, h))
        cw(ws, c, 15 if c == 1 else 13)

    for r, row in enumerate(util_df.itertuples(index=False), 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            fmt(cell)
            if c > 2 and isinstance(val, (int, float)):
                cell.number_format = "0.0%"
                if val > 1.0:
                    cell.fill = hf
                    cell.font = Font(name="Arial", size=9, bold=True,
                                     color=theme["high_font"])
                elif val >= 0.8:
                    cell.fill = mf
                    cell.font = Font(name="Arial", size=9, color=theme["medium_font"])
                else:
                    cell.fill = lf
                    cell.font = Font(name="Arial", size=9, color=theme["low_font"])

    leg = util_df.shape[0] + 6
    section_hdr(ws, leg, "COLOR LEGEND", 4)
    for i, (lbl, bg, fg) in enumerate([
        ("Overloaded  > 100%",     theme["high"],   theme["high_font"]),
        ("Near Capacity  80–100%", theme["medium"],  theme["medium_font"]),
        ("Healthy  < 80%",         theme["low"],    theme["low_font"]),
    ]):
        c = ws.cell(leg + 1 + i, 1, lbl)
        c.fill      = PatternFill("solid", start_color=bg)
        c.font      = Font(name="Arial", size=9, bold=True, color=fg)
        c.border    = bdr()
        c.alignment = Alignment(horizontal="left", vertical="center")

    cw(ws, 1, 24)
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"


def write_summary_sheet(wb, weekly_df, util_df, types, capacities, years,
                        tool_title, cap_note=""):
    ws = wb.create_sheet("Summary")
    banner(ws, 1, f"📊  {tool_title} — Executive Summary", cols=20)

    # Capacity info
    section_hdr(ws, 3, "⚙  Capacity Configuration", 6)
    hdr(ws.cell(4, 1, "Lab Type"),        bg="2E75B6", sz=9)
    hdr(ws.cell(4, 2, "Annual Capacity"), bg="2E75B6", sz=9)
    hdr(ws.cell(4, 3, "Weekly Capacity"), bg="2E75B6", sz=9)
    cw(ws, 1, 18); cw(ws, 2, 16); cw(ws, 3, 16)
    for i, (k, v) in enumerate(capacities.items(), 5):
        ws.cell(i, 1, k).font   = Font(name="Arial", size=9)
        ws.cell(i, 1).border    = bdr()
        ws.cell(i, 1).alignment = Alignment(horizontal="left")
        c2 = ws.cell(i, 2, v)
        c2.font = Font(name="Arial", size=9, bold=True); c2.border = bdr()
        c2.alignment = Alignment(horizontal="center")
        c3 = ws.cell(i, 3, round(v / 52, 2))
        c3.font = Font(name="Arial", size=9); c3.border = bdr()
        c3.number_format = "0.00"; c3.alignment = Alignment(horizontal="center")
    if cap_note:
        nr = 5 + len(capacities)
        ws.cell(nr, 1, f"ℹ  {cap_note}").font = Font(name="Arial", size=8,
                                                        italic=True, color="666666")

    # Annual demand
    s = 5 + len(capacities) + 3
    section_hdr(ws, s, "📈  Annual Demand by Lab Type (samples/year)", 2 + len(years))
    hdrs = ["Lab Type"] + [str(y) for y in years] + ["Average"]
    for c, h in enumerate(hdrs, 1):
        hdr(ws.cell(s + 1, c, h), bg="2E75B6", sz=9)
        cw(ws, c, 14)
    for ri, t in enumerate(types, s + 2):
        ws.cell(ri, 1, t).font   = Font(name="Arial", size=9, bold=True)
        ws.cell(ri, 1).border    = bdr()
        ws.cell(ri, 1).alignment = Alignment(horizontal="left")
        vals = []
        for ci, y in enumerate(years, 2):
            total = weekly_df[weekly_df["Year"] == y][t].sum()
            cell  = ws.cell(ri, ci, round(total, 1))
            fmt(cell, number_format="#,##0.0")
            vals.append(total)
        ac = ws.cell(ri, 2 + len(years), round(sum(vals) / len(vals), 1))
        fmt(ac, number_format="#,##0.0", bold=True)
        ac.fill = PatternFill("solid", start_color="D9E1F2")

    # Peak utilization
    s2 = s + 2 + len(types) + 2
    section_hdr(ws, s2, "🚦  Peak Weekly Utilization (worst week per year)", 2 + len(years))
    hdrs2 = ["Lab Type"] + [str(y) for y in years]
    for c, h in enumerate(hdrs2, 1):
        hdr(ws.cell(s2 + 1, c, h), bg="C00000", sz=9)
    for ri, t in enumerate(types, s2 + 2):
        ws.cell(ri, 1, t).font   = Font(name="Arial", size=9, bold=True)
        ws.cell(ri, 1).border    = bdr()
        ws.cell(ri, 1).alignment = Alignment(horizontal="left")
        for ci, y in enumerate(years, 2):
            peak = util_df[util_df["Year"] == y][t].max()
            cell = ws.cell(ri, ci, round(peak, 4))
            fmt(cell, number_format="0.0%")
            if peak > 1.0:
                cell.fill = PatternFill("solid", start_color="FF4444")
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            elif peak >= 0.8:
                cell.fill = PatternFill("solid", start_color="FFD700")
                cell.font = Font(name="Arial", size=9, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color="70AD47")
                cell.font = Font(name="Arial", size=9, color="FFFFFF")


def write_utilization_chart(wb, util_df, types, years):
    """Monthly utilization line charts — one per year."""
    ws = wb.create_sheet("Chart_Utilization")
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    banner(ws, 1, "Utilization Trend — Monthly Average", cols=30)
    BLOCK = 15
    DR0   = 3

    for yi, year in enumerate(years):
        bs = DR0 + yi * BLOCK
        # Year label
        ws.merge_cells(start_row=bs, start_column=1,
                       end_row=bs, end_column=len(types) + 1)
        c = ws.cell(bs, 1, f"  Year {year}")
        c.font = Font(bold=True, size=10, name="Arial", color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6")
        ws.row_dimensions[bs].height = 15

        # Header
        ws.cell(bs+1, 1, "Month").font = Font(bold=True, name="Arial", size=9)
        ws.cell(bs+1, 1).fill = PatternFill("solid", start_color="D9E1F2")
        for ci, t in enumerate(types, 2):
            h = ws.cell(bs+1, ci, t)
            h.font = Font(bold=True, name="Arial", size=9)
            h.fill = PatternFill("solid", start_color="D9E1F2")
            h.alignment = Alignment(horizontal="center")
            cw(ws, ci, 14)
        cw(ws, 1, 10)

        # Monthly data
        yd = util_df[util_df["Year"] == year].copy()
        yd["_m"] = yd["Week"].apply(week_to_month)
        for mi, month in enumerate(months):
            dr = bs + 2 + mi
            ws.cell(dr, 1, month).font = Font(bold=True, name="Arial", size=9)
            ws.cell(dr, 1).fill = PatternFill("solid", start_color="EBF3FB")
            for ci, t in enumerate(types, 2):
                grp = yd[yd["_m"] == mi + 1][t]
                avg = grp.mean() if not grp.empty else 0.0
                if math.isnan(avg): avg = 0.0
                cell = ws.cell(dr, ci, round(avg, 4))
                cell.number_format = "0.0%"
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center")

        # Line chart
        chart = LineChart()
        chart.title        = f"Utilization Trend — {year}"
        chart.y_axis.title = "Utilization %"
        chart.x_axis.title = "Month"
        chart.y_axis.numFmt = "0%"
        chart.y_axis.scaling.min = 0
        chart.style  = 10
        chart.height = 13
        chart.width  = 22

        for ci in range(2, 2 + len(types)):
            dr = Reference(ws, min_col=ci, max_col=ci,
                           min_row=bs+1, max_row=bs+1+12)
            chart.add_data(dr, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=bs+2, max_row=bs+1+12)
        chart.set_categories(cats)

        cc = get_column_letter(len(types) + 3)
        ws.add_chart(chart, f"{cc}{bs}")

    ws.freeze_panes = "B3"


def write_capacity_chart(wb, weekly_df, types, capacities, years):
    """Capacity vs Avg & Peak demand — matches reference model exactly."""
    ws = wb.create_sheet("Chart_Capacity_vs_Demand")
    banner(ws, 1, "Average Weekly Demand vs Capacity — All Years", cols=30)

    HR = 3; D0 = 4; DE = D0 + len(types) - 1

    # Columns: LabType | Capacity | [yr Avg] x N | [yr Peak] x N
    avg_cols  = {}
    peak_cols = {}
    cw(ws, 1, 16); cw(ws, 2, 12)
    hdr(ws.cell(HR, 1, "Lab Type"),  bg="1F3864", sz=9)
    hdr(ws.cell(HR, 2, "Capacity"),  bg="1F3864", sz=9)
    for i, y in enumerate(years):
        ac = 3 + i
        hdr(ws.cell(HR, ac, f"{y} Avg Demand"),  bg="2E75B6", sz=9); cw(ws, ac, 14)
        avg_cols[y] = ac
    for i, y in enumerate(years):
        pc = 3 + len(years) + i
        hdr(ws.cell(HR, pc, f"{y} Peak Demand"), bg="1F5C1A", sz=9); cw(ws, pc, 14)
        peak_cols[y] = pc

    for ri, t in enumerate(types, D0):
        cap = capacities.get(t, 0)
        cap_wk = cap / 52 if cap > 0 else 0
        c1 = ws.cell(ri, 1, t); c1.font = Font(name="Arial", size=9, bold=True)
        c1.border = bdr(); c1.alignment = Alignment(horizontal="left")
        c2 = ws.cell(ri, 2, round(cap_wk, 2))
        fmt(c2, number_format="0.00")
        c2.fill = PatternFill("solid", start_color="EBF3FB")
        c2.font = Font(name="Arial", size=9, bold=True, color="1F3864")
        for y in years:
            yd  = weekly_df[weekly_df["Year"] == y]
            avg  = round(yd[t].mean(), 2) if t in yd.columns else 0
            peak = round(yd[t].max(), 2)  if t in yd.columns else 0
            util_avg  = avg  / cap_wk if cap_wk > 0 else 0
            util_peak = peak / cap_wk if cap_wk > 0 else 0
            # Avg cell
            ca = ws.cell(ri, avg_cols[y], avg); fmt(ca, number_format="0.00")
            ca.fill = PatternFill("solid", start_color=("FF4444" if util_avg>1 else "FFD700" if util_avg>=0.8 else "C6EFCE"))
            if util_avg > 1: ca.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            # Peak cell
            cp = ws.cell(ri, peak_cols[y], peak); fmt(cp, number_format="0.00")
            cp.fill = PatternFill("solid", start_color=("FF4444" if util_peak>1 else "FFD700" if util_peak>=0.8 else "C6EFCE"))
            if util_peak > 1: cp.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    # Bar chart — weekly capacity + avg demand per year
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Average Weekly Demand vs Capacity"
    chart.y_axis.title = "Weekly Units"; chart.x_axis.title = "Lab Type"
    chart.style = 10; chart.height = 16; chart.width = 32
    cats = Reference(ws, min_col=1, min_row=D0, max_row=DE)
    cap_ref = Reference(ws, min_col=2, max_col=2, min_row=HR, max_row=DE)
    chart.add_data(cap_ref, titles_from_data=True); chart.set_categories(cats)
    chart.series[0].title = SeriesLabel(v="Weekly Capacity")
    for y in years:
        ref = Reference(ws, min_col=avg_cols[y], max_col=avg_cols[y], min_row=HR, max_row=DE)
        chart.add_data(ref, titles_from_data=True)
    ws.add_chart(chart, f"A{DE + 4}")


def write_yoy_chart(wb, weekly_df, types, years):
    """YoY annual demand grouped bar chart."""
    ws = wb.create_sheet("Chart_YoY")
    banner(ws, 1, "Year-on-Year Demand Comparison", cols=20)

    HR = 3; D0 = 4; DE = 3 + len(types)
    hdrs = ["Lab Type"] + [str(y) for y in years] + ["Growth"]
    for c, h in enumerate(hdrs, 1):
        hdr(ws.cell(HR, c, h), bg="2E75B6", sz=9)
        cw(ws, c, 14)

    for ri, t in enumerate(types, D0):
        ws.cell(ri, 1, t).font   = Font(name="Arial", size=9, bold=True)
        ws.cell(ri, 1).border    = bdr()
        ws.cell(ri, 1).alignment = Alignment(horizontal="left")
        ytot = []
        for ci, y in enumerate(years, 2):
            total = weekly_df[weekly_df["Year"] == y][t].sum()
            cell  = ws.cell(ri, ci, round(total, 1))
            fmt(cell, number_format="#,##0.0")
            ytot.append(total)
        gc = ws.cell(ri, 2 + len(years))
        if len(ytot) >= 2 and ytot[0] > 0:
            g = (ytot[-1] - ytot[0]) / ytot[0]
            gc.value = round(g, 4); gc.number_format = "+0.0%;-0.0%;0.0%"
            gc.font  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            gc.fill  = PatternFill("solid",
                                    start_color="70AD47" if g >= 0 else "FF4444")
            gc.border = bdr(); gc.alignment = Alignment(horizontal="center")
        else:
            gc.value = "N/A"; fmt(gc)

    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Year-on-Year Annual Demand"; chart.style = 10
    chart.y_axis.title = "Samples/Year"; chart.x_axis.title = "Lab Type"
    chart.height = 15; chart.width = 28

    cats = Reference(ws, min_col=1, min_row=D0, max_row=DE)
    first = Reference(ws, min_col=2, max_col=2, min_row=HR, max_row=DE)
    chart.add_data(first, titles_from_data=True)
    chart.set_categories(cats)
    for ci in range(3, 2 + len(years)):
        ref = Reference(ws, min_col=ci, max_col=ci, min_row=HR, max_row=DE)
        chart.add_data(ref, titles_from_data=True)
    ws.add_chart(chart, f"A{DE + 4}")

    # ── PIE CHARTS per year (process-wise demand split) ──
    from openpyxl.chart import PieChart, ProjectedPieChart
    pie_anchor_col = get_column_letter(3 + len(years) + 2)
    pie_row = 3

    ws.cell(pie_row - 1, 3 + len(years) + 2,
            "Process-wise Demand Share (%) by Year").font = Font(
                bold=True, size=10, name="Arial", color="1F3864")

    for yi, year in enumerate(years):
        # Build a small data table for this year's pie
        pie_data_col  = 3 + len(years) + 2 + yi * 3
        pie_label_col = pie_data_col - 1

        # Labels + values in two hidden columns
        ws.cell(pie_row, pie_label_col, str(year)).font = Font(bold=True, name="Arial", size=9)
        for ri2, t in enumerate(types, pie_row + 1):
            ws.cell(ri2, pie_label_col, t)
            total_val = weekly_df[weekly_df["Year"] == year][t].sum()
            ws.cell(ri2, pie_data_col, round(total_val, 1))

        pie = PieChart()
        pie.title  = str(year)
        pie.style  = 10
        pie.height = 12
        pie.width  = 14

        data_ref  = Reference(ws,
                              min_col=pie_data_col, max_col=pie_data_col,
                              min_row=pie_row, max_row=pie_row + len(types))
        label_ref = Reference(ws,
                              min_col=pie_label_col,
                              min_row=pie_row + 1, max_row=pie_row + len(types))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(label_ref)
        pie.dataLabels           = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal     = False

        anchor_col = get_column_letter(pie_label_col)
        anchor_row = pie_row + len(types) + 2
        ws.add_chart(pie, f"{anchor_col}{anchor_row}")


def write_gantt_current_year(wb, weekly_df, types, capacities,
                              current_year, current_week):
    """
    Gantt occupancy chart for current year only.
    Shows weeks 1..current_week (actual data) clearly marked.
    Skips if data is too sparse (< 5 samples/week average).
    """
    ws = wb.create_sheet("Gantt_Current_Year")

    # Filter current year
    yd = weekly_df[weekly_df["Year"] == current_year]
    if yd.empty:
        ws.cell(1, 1, f"No data for {current_year}").font = \
            Font(bold=True, name="Arial", size=11, color="FF0000")
        return

    # Check if data is sufficient
    total_demand = sum(yd[t].sum() for t in types)
    if total_demand < 0.5:
        ws.cell(1, 1,
            f"Insufficient data for {current_year} Gantt chart "
            f"(total weekly demand < 0.5). "
            f"Please ensure {current_year} data is present in the input file."
        ).font = Font(bold=True, name="Arial", size=10, color="FF4444")
        return

    banner(ws, 1,
           f"🗓  Gantt — Lab Occupancy {current_year}  "
           f"(Current Week: {current_week})",
           cols=55, height=28)

    # Color fills
    red_f  = PatternFill("solid", start_color="FF4444")
    yel_f  = PatternFill("solid", start_color="FFD700")
    grn_f  = PatternFill("solid", start_color="70AD47")
    grey_f = PatternFill("solid", start_color="F2F2F2")
    fut_f  = PatternFill("solid", start_color="ECECEC")  # future weeks

    # Header row
    HR = 3
    hdr(ws.cell(HR, 1, "Lab Type"), bg="1F3864", sz=9); cw(ws, 1, 16)
    hdr(ws.cell(HR, 2, "Cap/Wk"),   bg="1F3864", sz=9); cw(ws, 2, 7)
    for w in range(1, 53):
        c = ws.cell(HR, w + 2, w)
        if w == current_week:
            c.fill = PatternFill("solid", start_color="C00000")
            c.font = Font(bold=True, size=7, name="Arial", color="FFFFFF")
        elif w > current_week:
            c.fill = PatternFill("solid", start_color="D9D9D9")
            c.font = Font(size=6, name="Arial", color="AAAAAA")
        else:
            c.fill = PatternFill("solid", start_color="2E75B6")
            c.font = Font(bold=True, size=7, name="Arial", color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        cw(ws, w + 2, 1.8)
    ws.row_dimensions[HR].height = 14

    # Quarter markers above header
    QR = 2
    for label, start_col, end_col, bg in [
        ("Q1: Wk 1–13",  3,  15, "4472C4"),
        ("Q2: Wk 14–26", 16, 28, "2E75B6"),
        ("Q3: Wk 27–39", 29, 41, "1F4E79"),
        ("Q4: Wk 40–52", 42, 54, "264478"),
    ]:
        ws.merge_cells(start_row=QR, start_column=start_col,
                       end_row=QR, end_column=end_col)
        c = ws.cell(QR, start_col, label)
        c.font      = Font(bold=True, size=8, name="Arial", color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[QR].height = 13

    # Lab rows
    for ri, t in enumerate(types, HR + 1):
        cap_wk = capacities.get(t, 1) / 52
        lc = ws.cell(ri, 1, t)
        lc.font      = Font(bold=True, size=9, name="Arial")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = bdr()

        cc = ws.cell(ri, 2, round(cap_wk, 2))
        cc.font      = Font(size=8, name="Arial")
        cc.alignment = Alignment(horizontal="center")
        cc.border    = bdr()

        for w in range(1, 53):
            wrow = yd[yd["Week"] == w]
            dem  = wrow[t].values[0] if not wrow.empty else 0.0
            util = dem / cap_wk if cap_wk > 0 else 0.0
            cell = ws.cell(ri, w + 2)
            cell.border = bdr("EEEEEE")

            if w > current_week:
                cell.fill  = fut_f
                cell.value = ""
            elif util > 1.0:
                cell.fill  = red_f
                cell.value = f"{util:.0%}"
                cell.font  = Font(size=5, name="Arial", color="FFFFFF", bold=True)
            elif util >= 0.8:
                cell.fill  = yel_f
                cell.value = f"{util:.0%}"
                cell.font  = Font(size=5, name="Arial", color="333333")
            elif util >= 0.3:
                cell.fill  = grn_f
                cell.value = f"{util:.0%}"
                cell.font  = Font(size=5, name="Arial", color="FFFFFF")
            else:
                cell.fill  = grey_f
                cell.value = ""

            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 14

    # Legend
    leg = HR + len(types) + 3
    section_hdr(ws, leg, "LEGEND", 8, bg="2E75B6")
    for i, (lbl, bg, fg) in enumerate([
        ("Overloaded  > 100%",     "FF4444", "FFFFFF"),
        ("Near Capacity  80–100%", "FFD700", "333333"),
        ("Active  30–80%",         "70AD47", "FFFFFF"),
        ("Idle  < 30%",            "F2F2F2", "666666"),
        ("Future weeks (no data)", "ECECEC", "AAAAAA"),
    ]):
        c = ws.cell(leg + 1 + i, 1, lbl)
        c.fill      = PatternFill("solid", start_color=bg)
        c.font      = Font(name="Arial", size=9, bold=True, color=fg)
        c.border    = bdr()
        c.alignment = Alignment(horizontal="left")
    ws.cell(leg + 6, 1,
            f"▼ Red column = Current Week ({current_week})").font = \
        Font(name="Arial", size=8, italic=True, color="C00000")

    ws.freeze_panes = "C4"


def write_gantt_all_years(wb, weekly_df, types, capacities, years):
    """
    Lab Occupancy Gantt — Excel Project Planner style (matches reference template).

    Left summary columns (Lab Type | Year | Ann. Cap | Ann. Demand | Util % | Status)
    then 52 week columns colour-coded:
      GREEN  = within capacity (<80%)
      YELLOW = near capacity (80-100%)
      RED    = over capacity (>100%)
      GREY   = vacant
    Current week column = tan fill + orange header (Period Highlight).
    """
    ws = wb.create_sheet("Gantt_Chart")
    banner(ws, 1, "Lab Occupancy Gantt — Project Planner Style (Lab Type × Year × Week)", cols=60)

    GREY_BG, GREEN_BG, YELLOW_BG, RED_BG = "D9D9D9","70AD47","FFD700","FF4444"
    HL_BG, HL_HDR = "FDE9D9","C55A11"   # tan fill, orange header

    # Legend row
    LR = 2
    ws.cell(LR, 1, "Legend:").font = Font(bold=True, size=9, name="Arial")
    legend_items = [
        ("Within capacity",             GREEN_BG,  "FFFFFF", 9),
        ("Near capacity (80-100%)",     YELLOW_BG, "333333", 13),
        ("Over capacity",               RED_BG,    "FFFFFF", 10),
        ("Vacant / no demand",          GREY_BG,   "666666", 10),
        ("Period Highlight (cur. week)",HL_BG,     "C55A11",  8),
    ]
    col = 2
    for lbl, bg, fg, span in legend_items:
        ws.merge_cells(start_row=LR, start_column=col, end_row=LR, end_column=col+span-1)
        c = ws.cell(LR, col, lbl)
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(name="Arial", size=8, bold=True, color=fg)
        c.alignment = Alignment(horizontal="center")
        col += span + 1
    ws.row_dimensions[LR].height = 14

    QR, HR = 3, 4
    LEFT_HDRS   = ["Lab Type","Year","Ann. Cap","Ann. Demand","Util %","Status"]
    LEFT_WIDTHS = [14, 7, 10, 12, 8, 12]
    N_LEFT = len(LEFT_HDRS)

    for ci, (h, w) in enumerate(zip(LEFT_HDRS, LEFT_WIDTHS), 1):
        hdr(ws.cell(HR, ci, h), bg="1F3864", sz=8); cw(ws, ci, w)

    # Quarter row
    for q_lbl, q_s, q_e in [("Q1",1,13),("Q2",14,26),("Q3",27,39),("Q4",40,52)]:
        sc = N_LEFT + q_s; ec = N_LEFT + q_e
        ws.merge_cells(start_row=QR, start_column=sc, end_row=QR, end_column=ec)
        c = ws.cell(QR, sc, q_lbl)
        c.font = Font(bold=True, size=8, name="Arial", color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6")
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[QR].height = 13

    hl_week   = CURRENT_WEEK if CURRENT_YEAR in [int(y) for y in years] else None
    thick_sd  = Side(style="medium", color=HL_HDR)

    for w in range(1, 53):
        c = ws.cell(HR, N_LEFT + w, w)
        c.alignment = Alignment(horizontal="center")
        cw(ws, N_LEFT + w, 1.8)
        if hl_week == w:
            c.fill = PatternFill("solid", start_color=HL_HDR)
            c.font = Font(bold=True, size=7, name="Arial", color="FFFFFF")
            c.border = Border(left=thick_sd, right=thick_sd, top=thick_sd, bottom=thick_sd)
        else:
            c.fill = PatternFill("solid", start_color="2E75B6")
            c.font = Font(bold=True, size=7, name="Arial", color="FFFFFF")
    ws.row_dimensions[HR].height = 14

    ri = HR + 1
    for t in types:
        for year in sorted(years):
            yd       = weekly_df[weekly_df["Year"] == year]
            cap_yr   = capacities.get(t, 1)
            cap_wk   = cap_yr / 52
            dem_yr   = yd[t].sum() if not yd.empty and t in yd.columns else 0.0
            util_pct = dem_yr / cap_yr * 100 if cap_yr > 0 else 0.0

            if util_pct > 100:
                st_str, st_bg, st_fg = "OVER CAPACITY", RED_BG, "FFFFFF"
            elif util_pct >= 80:
                st_str, st_bg, st_fg = "HIGH",          YELLOW_BG, "333333"
            else:
                st_str, st_bg, st_fg = "OK",            GREEN_BG, "FFFFFF"

            for ci, val in enumerate([t, year, int(cap_yr), int(round(dem_yr)),
                                        f"{util_pct:.1f}%", st_str], 1):
                c = ws.cell(ri, ci, val)
                c.font = Font(bold=(ci==1), size=9, name="Arial")
                c.border = bdr()
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left")
                if ci in (5, 6):
                    c.fill = PatternFill("solid", start_color=st_bg)
                    c.font = Font(bold=True, size=9, name="Arial", color=st_fg)

            for w in range(1, 53):
                wrow = yd[yd["Week"] == w]
                dem  = float(wrow[t].values[0]) if not wrow.empty else 0.0
                util = dem / cap_wk if cap_wk > 0 else 0.0

                if util > 1.0:
                    bg = RED_BG
                elif util >= 0.8:
                    bg = YELLOW_BG
                elif util >= 0.05:
                    bg = GREEN_BG
                else:
                    bg = GREY_BG

                col_idx = N_LEFT + w
                cell    = ws.cell(ri, col_idx)
                cell.fill      = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center")

                if hl_week == w and int(year) == CURRENT_YEAR:
                    cell.fill   = PatternFill("solid", start_color=HL_BG)
                    cell.border = Border(left=thick_sd, right=thick_sd,
                                          top=thick_sd, bottom=thick_sd)
                else:
                    cell.border = bdr("EEEEEE")

            ws.row_dimensions[ri].height = 12
            ri += 1

        ws.row_dimensions[ri].height = 5
        ri += 1

    ws.freeze_panes = f"G{HR+1}"


def write_gantt_heatmap(wb, weekly_df, types, capacities, years):
    """Numeric heatmap per year — matches reference Gantt_Heatmap sheet."""
    ws = wb.create_sheet("Gantt_Heatmap")
    banner(ws, 1, "Gantt Heatmap — Weekly Demand by Lab Type & Year (all values shown)", cols=55)

    red_f  = PatternFill("solid", start_color="FF4444")
    yel_f  = PatternFill("solid", start_color="FFD700")
    grn_f  = PatternFill("solid", start_color="C6EFCE")
    wht_f  = PatternFill("solid", start_color="FFFFFF")

    block_start = 3
    for year in sorted(years):
        yd = weekly_df[weekly_df["Year"]==year]
        # Year label
        ws.merge_cells(start_row=block_start, start_column=1,
                       end_row=block_start, end_column=53)
        yc = ws.cell(block_start, 1, f"  Year {year}")
        yc.font = Font(bold=True,size=10,name="Arial",color="FFFFFF")
        yc.fill = PatternFill("solid",start_color="2E75B6")
        yc.alignment = Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[block_start].height = 16

        # Header
        HR = block_start+1
        h0 = ws.cell(HR,1,"Lab / Week")
        h0.font=Font(bold=True,size=8,name="Arial"); h0.fill=PatternFill("solid",start_color="D9E1F2")
        h0.alignment=Alignment(horizontal="center"); h0.border=bdr(); cw(ws,1,14)
        for w in range(1,53):
            c=ws.cell(HR,w+1,w)
            c.font=Font(bold=True,size=7,name="Arial")
            c.fill=PatternFill("solid",start_color="D9E1F2")
            c.alignment=Alignment(horizontal="center"); c.border=bdr()
            cw(ws,w+1,2.5)

        # Data rows
        for ri2, t in enumerate(types, HR+1):
            cap_wk = capacities.get(t,1)/52
            lc=ws.cell(ri2,1,t)
            lc.font=Font(bold=True,size=9,name="Arial"); lc.border=bdr()
            lc.alignment=Alignment(horizontal="left")
            for w in range(1,53):
                wrow=yd[yd["Week"]==w]
                dem=float(wrow[t].values[0]) if not wrow.empty else 0.0
                util=dem/cap_wk if cap_wk>0 else 0.0
                cell=ws.cell(ri2,w+1,round(dem,2))
                cell.number_format="0.00"; cell.border=bdr()
                cell.alignment=Alignment(horizontal="center")
                cell.font=Font(size=8,name="Arial")
                if util>1.0:
                    cell.fill=red_f; cell.font=Font(size=8,name="Arial",bold=True,color="FFFFFF")
                elif util>=0.8:
                    cell.fill=yel_f
                elif util>=0.3:
                    cell.fill=grn_f
                else:
                    cell.fill=wht_f
            ws.row_dimensions[ri2].height=13

        block_start += 2 + len(types) + 1  # year label + header + rows + gap
    ws.freeze_panes = "B3"


def save_workbook(wb, input_path, filename):
    """Save to Lab_Dashboard_Output folder next to input file."""
    out_dir  = os.path.join(os.path.dirname(os.path.abspath(input_path)),
                            "Lab_Dashboard_Output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)
    return out_path
