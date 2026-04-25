"""
Tool 2 — Coating Labs Dashboard (Cold Spray, HVOF, Plasma)
============================================================
Combined capacity = 350 samples/year across all 3 coating labs.
Individual + combined utilization tracking.
Run: python tool2_coating_labs.py
"""

import os, sys, threading, math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import lab_core as core

# ── Tool-specific config ──────────────────────────────────
TOOL_TITLE    = "Coating Labs Dashboard (Cold Spray, HVOF, Plasma)"
ALLOWED_TYPES = ["Cold Spray", "HVOF", "Plasma"]
OUTPUT_FILE   = "Coating_Labs_Dashboard.xlsx"

# Each lab has its OWN independent capacity of 350 samples/year
DEFAULT_CAPS = {
    "Cold Spray": 350,
    "HVOF":       350,
    "Plasma":     350,
}


# ═══════════════════════════════════════════════════════════
#  COMBINED UTILIZATION SHEET
# ═══════════════════════════════════════════════════════════

def write_combined_sheet(wb, weekly_df, types, years, combined_cap):
    """Shows total combined demand across all coating labs vs 350 cap."""
    ws = wb.create_sheet("Combined_Utilization")
    core.banner(ws, 1,
                f"Combined Coating Labs — Total Demand vs {combined_cap}/yr Capacity",
                cols=20)

    weekly_cap = combined_cap / 52

    # Build combined weekly
    headers = ["Year", "Week", "Total Demand", "Capacity/Wk", "Utilization %",
               "Status"] + types
    for c, h in enumerate(headers, 1):
        core.hdr(ws.cell(3, c, h), bg="1F3864", sz=9)
        core.cw(ws, c, 14)
    core.cw(ws, 1, 8); core.cw(ws, 2, 7); core.cw(ws, 3, 14)
    core.cw(ws, 4, 12); core.cw(ws, 5, 14); core.cw(ws, 6, 12)

    red_f  = PatternFill("solid", start_color="FF4444")
    yel_f  = PatternFill("solid", start_color="FFD700")
    grn_f  = PatternFill("solid", start_color="70AD47")

    for r, row in enumerate(weekly_df.itertuples(index=False), 4):
        vals     = list(row)
        year_v   = vals[0]; week_v = vals[1]
        type_vals = {t: vals[2 + i] for i, t in enumerate(types)}
        total    = sum(type_vals.values())
        util     = total / weekly_cap if weekly_cap > 0 else 0.0

        ws.cell(r, 1, year_v); core.fmt(ws.cell(r, 1))
        ws.cell(r, 2, week_v); core.fmt(ws.cell(r, 2))

        tc = ws.cell(r, 3, round(total, 3)); core.fmt(tc, number_format="0.000")
        if util > 1.0:   tc.fill = red_f;  tc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        elif util >= 0.8: tc.fill = yel_f
        else:             tc.fill = grn_f;  tc.font = Font(name="Arial", size=9, color="FFFFFF")

        ws.cell(r, 4, round(weekly_cap, 2)); core.fmt(ws.cell(r, 4), number_format="0.00")

        uc = ws.cell(r, 5, round(util, 4)); core.fmt(uc, number_format="0.0%")
        if util > 1.0:   uc.fill = red_f;  uc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        elif util >= 0.8: uc.fill = yel_f
        else:             uc.fill = grn_f;  uc.font = Font(name="Arial", size=9, color="FFFFFF")

        status = "🔴 OVERLOAD" if util > 1.0 else "🟡 Near Cap" if util >= 0.8 else "🟢 OK"
        ws.cell(r, 6, status); core.fmt(ws.cell(r, 6))

        for ci, t in enumerate(types, 7):
            ws.cell(r, ci, round(type_vals[t], 3)); core.fmt(ws.cell(r, ci), number_format="0.000")

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"


def write_combined_chart(wb, weekly_df, types, years, combined_cap):
    """Line chart: combined weekly demand vs combined capacity."""
    ws = wb.create_sheet("Chart_Combined")
    core.banner(ws, 1,
                f"Combined Coating Labs — Weekly Demand Trend vs {combined_cap}/yr Cap",
                cols=30)

    # Build monthly aggregated combined demand
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    weekly_cap = combined_cap / 52

    HR = 3
    ws.cell(HR, 1, "Month").font = Font(bold=True, name="Arial", size=9)
    ws.cell(HR, 1).fill = PatternFill("solid", start_color="D9E1F2")

    yr_cols = {}
    for i, y in enumerate(years):
        col = 2 + i
        h = ws.cell(HR, col, str(y))
        h.font = Font(bold=True, name="Arial", size=9)
        h.fill = PatternFill("solid", start_color="D9E1F2")
        h.alignment = Alignment(horizontal="center")
        core.cw(ws, col, 14)
        yr_cols[y] = col
    core.cw(ws, 1, 10)

    # Capacity line column
    cap_col = 2 + len(years)
    hc = ws.cell(HR, cap_col, f"Cap/Wk ({combined_cap}/yr)")
    hc.font = Font(bold=True, name="Arial", size=9)
    hc.fill = PatternFill("solid", start_color="FFD700")
    hc.alignment = Alignment(horizontal="center")
    core.cw(ws, cap_col, 16)

    for mi, month in enumerate(months):
        row = HR + 1 + mi
        ws.cell(row, 1, month).font = Font(bold=True, name="Arial", size=9)
        ws.cell(row, 1).fill = PatternFill("solid", start_color="EBF3FB")

        for y in years:
            yd = weekly_df[weekly_df["Year"] == y].copy()
            yd["_m"] = yd["Week"].apply(core.week_to_month)
            month_rows = yd[yd["_m"] == mi + 1]
            total_avg  = sum(month_rows[t].mean() for t in types
                             if not month_rows.empty)
            if math.isnan(total_avg): total_avg = 0.0
            cell = ws.cell(row, yr_cols[y], round(total_avg, 3))
            cell.number_format = "0.000"
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center")

        # Capacity reference
        cc = ws.cell(row, cap_col, round(weekly_cap, 2))
        cc.number_format = "0.00"
        cc.font      = Font(name="Arial", size=9, bold=True)
        cc.fill      = PatternFill("solid", start_color="FFF2CC")
        cc.alignment = Alignment(horizontal="center")

    # Line chart
    chart = LineChart()
    chart.title        = f"Combined Coating Labs Weekly Demand (Cap={combined_cap}/yr)"
    chart.y_axis.title = "Weekly Demand (samples)"
    chart.x_axis.title = "Month"
    chart.style        = 10
    chart.height       = 16
    chart.width        = 30

    # Add each year series
    first_col = min(yr_cols.values())
    first_ref = Reference(ws, min_col=first_col, max_col=first_col,
                          min_row=HR, max_row=HR + 12)
    chart.add_data(first_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=HR+1, max_row=HR+12)
    chart.set_categories(cats)

    for y in list(years)[1:]:
        col = yr_cols[y]
        ref = Reference(ws, min_col=col, max_col=col, min_row=HR, max_row=HR+12)
        chart.add_data(ref, titles_from_data=True)

    # Capacity reference line
    cap_ref = Reference(ws, min_col=cap_col, max_col=cap_col,
                        min_row=HR, max_row=HR+12)
    chart.add_data(cap_ref, titles_from_data=True)

    ws.add_chart(chart, f"A{HR + 15}")

    # Also add stacked bar chart per type
    chart2 = BarChart()
    chart2.type     = "col"
    chart2.grouping = "stacked"
    chart2.title    = "Demand by Type — Stacked (Monthly)"
    chart2.y_axis.title = "Weekly Demand"
    chart2.x_axis.title = "Month"
    chart2.style    = 10
    chart2.height   = 16
    chart2.width    = 28

    # Write per-type monthly data for last year
    last_yr = max(years)
    type_start_col = cap_col + 2
    ws.cell(HR, type_start_col - 1, f"Per-Type {last_yr}").font = \
        Font(bold=True, name="Arial", size=9)
    for ti, t in enumerate(types):
        col = type_start_col + ti
        h2  = ws.cell(HR, col, t)
        h2.font = Font(bold=True, name="Arial", size=9)
        h2.fill = PatternFill("solid", start_color="D9E1F2")
        h2.alignment = Alignment(horizontal="center")
        core.cw(ws, col, 13)
        yd = weekly_df[weekly_df["Year"] == last_yr].copy()
        yd["_m"] = yd["Week"].apply(core.week_to_month)
        for mi in range(12):
            row  = HR + 1 + mi
            grp  = yd[yd["_m"] == mi + 1][t]
            avg  = grp.mean() if not grp.empty else 0.0
            cell = ws.cell(row, col, round(avg if not math.isnan(avg) else 0.0, 3))
            cell.number_format = "0.000"
            cell.alignment     = Alignment(horizontal="center")
            cell.font          = Font(name="Arial", size=9)

    for ti, t in enumerate(types):
        col  = type_start_col + ti
        ref2 = Reference(ws, min_col=col, max_col=col, min_row=HR, max_row=HR+12)
        chart2.add_data(ref2, titles_from_data=True)
    chart2.set_categories(cats)
    chart2_col = get_column_letter(type_start_col)
    ws.add_chart(chart2, f"{chart2_col}{HR + 15}")


# ═══════════════════════════════════════════════════════════
#  MAIN GENERATOR
# ═══════════════════════════════════════════════════════════

def generate(input_path, individual_caps, theme, progress_cb=None):
    def p(msg):
        if progress_cb: progress_cb(msg)

    p("📖 Reading file…")
    df, col_map, errors, warns = core.load_and_filter(input_path, ALLOWED_TYPES)
    if errors:
        raise ValueError("\n".join(errors))

    p("⚙️  Computing weekly demand…")
    weekly_df, util_df, types, years = core.build_weekly(df, col_map, individual_caps)

    p("📗 Building workbook…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    p("📄 Sheet 1/7 — Weekly Planner…")
    core.write_weekly_planner(wb, weekly_df, types, TOOL_TITLE)

    p("🎨 Sheet 2/7 — Utilization…")
    core.write_utilization_sheet(wb, util_df, types, theme)

    p("📋 Sheet 3/7 — Summary…")
    core.write_summary_sheet(wb, weekly_df, util_df, types, individual_caps, years,
                              TOOL_TITLE,
                              cap_note="  ".join(f"{k}={v}/yr" for k, v in individual_caps.items()))

    p("📈 Sheet 4/7 — Utilization Chart…")
    core.write_utilization_chart(wb, util_df, types, years)

    p("📊 Sheet 5/7 — Capacity vs Demand…")
    core.write_capacity_chart(wb, weekly_df, types, individual_caps, years)

    p("📊 Sheet 6/7 — Year-on-Year…")
    core.write_yoy_chart(wb, weekly_df, types, years)

    # Gantt current year
    current_year = core.CURRENT_YEAR
    current_week = core.CURRENT_WEEK
    gantt_year   = current_year if current_year in years else max(years)
    gantt_week   = current_week if gantt_year == current_year else 52
    p(f"🗓️  Sheet 7/7 — Gantt {gantt_year}…")
    core.write_gantt_current_year(wb, weekly_df, types, individual_caps,
                                   gantt_year, gantt_week)

    p("💾 Saving…")
    out = core.save_workbook(wb, input_path, OUTPUT_FILE)
    return out, warns


# ═══════════════════════════════════════════════════════════
#  GUI
# ═══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"🏭  {TOOL_TITLE}")
        self.geometry("740x620")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self.file_path    = tk.StringVar(value="No file selected")
        self.ind_caps     = {k: tk.IntVar(value=v) for k, v in DEFAULT_CAPS.items()}
        self.theme_var    = tk.StringVar(value=list(core.COLOR_THEMES.keys())[0])
        self.progress_var = tk.StringVar(value="")
        self._build()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _build(self):
        hdr = tk.Frame(self, bg="#1F5C1A", height=72)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🏭  Coating Labs Dashboard",
                 bg="#1F5C1A", fg="white",
                 font=("Arial", 14, "bold")).pack(pady=8)
        tk.Label(hdr, text="Cold Spray  |  HVOF  |  Plasma   —   Each lab: 350 samples/yr (independent)",
                 bg="#1F5C1A", fg="#C6EFCE",
                 font=("Arial", 9)).pack()

        main = tk.Frame(self, bg="#F0F4F8", padx=20, pady=10)
        main.pack(fill="both", expand=True)

        # File
        self._sec(main, "📂 Step 1 — Select Input Excel File")
        ff = tk.Frame(main, bg="#F0F4F8"); ff.pack(fill="x", pady=(0,8))
        self._btn(ff, "📁 Browse…", self._pick).pack(side="left")
        tk.Label(ff, textvariable=self.file_path, bg="#F0F4F8", fg="#444",
                 font=("Arial", 9), wraplength=520).pack(side="left", padx=10)

        ttk.Separator(main).pack(fill="x", pady=6)

        # Individual caps only — each lab is independent
        self._sec(main, "⚙️ Step 2 — Individual Lab Capacities (samples/year — each lab independent)")
        icf = tk.Frame(main, bg="#F0F4F8"); icf.pack(fill="x", pady=(0,8))
        for i, (name, var) in enumerate(self.ind_caps.items()):
            card2 = tk.Frame(icf, bg="#FFFFFF", bd=1, relief="groove", padx=12, pady=8)
            card2.grid(row=0, column=i, padx=6, sticky="ew")
            icf.columnconfigure(i, weight=1)
            tk.Label(card2, text=name, bg="#FFFFFF",
                     font=("Arial", 10, "bold"), fg="#1F5C1A").pack(anchor="w")
            tk.Label(card2, text="samples/year (independent)", bg="#FFFFFF",
                     font=("Arial", 8), fg="#666").pack(anchor="w")
            tk.Spinbox(card2, from_=1, to=9999, textvariable=var,
                       width=10, font=("Arial", 11),
                       relief="flat", bd=1).pack(anchor="w", pady=(3,0))

        ttk.Separator(main).pack(fill="x", pady=6)

        # Theme
        self._sec(main, "🎨 Step 3 — Color Theme")
        tf = tk.Frame(main, bg="#F0F4F8"); tf.pack(fill="x", pady=(0,8))
        for i, name in enumerate(core.COLOR_THEMES):
            tk.Radiobutton(tf, text=name, variable=self.theme_var,
                           value=name, bg="#F0F4F8",
                           font=("Arial", 9)).grid(row=0, column=i, padx=8, sticky="w")

        ttk.Separator(main).pack(fill="x", pady=6)

        self.prog_label = tk.Label(main, textvariable=self.progress_var,
                                   bg="#F0F4F8", fg="#1F5C1A",
                                   font=("Arial", 9, "italic"))
        self.prog_label.pack(anchor="w")
        self.prog_bar = ttk.Progressbar(main, mode="indeterminate", length=680)
        self.prog_bar.pack(fill="x", pady=4)

        self.gen_btn = self._btn(main, "⚡ Generate Coating Labs Dashboard",
                                 self._run, color="#1F5C1A", size=11, pady=13)
        self.gen_btn.pack(pady=10, fill="x")

        ftr = tk.Frame(self, bg="#DDE5EF", height=26)
        ftr.pack(fill="x", side="bottom")
        tk.Label(ftr, text=f"Output → Lab_Dashboard_Output/{OUTPUT_FILE}",
                 bg="#DDE5EF", fg="#666", font=("Arial", 8)).pack(pady=5)

    def _sec(self, p, title):
        tk.Label(p, text=title, bg="#F0F4F8",
                 font=("Arial", 10, "bold"), fg="#1F5C1A").pack(anchor="w", pady=(6,2))

    def _btn(self, p, text, cmd, color="#1F5C1A", size=9, pady=8):
        return tk.Button(p, text=text, command=cmd,
                         bg=color, fg="white", activebackground=color,
                         font=("Arial", size, "bold"), relief="flat",
                         cursor="hand2", padx=16, pady=pady)

    def _pick(self):
        path = filedialog.askopenfilename(
            title="Select Lab Data Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if path:
            self.file_path.set(path)
            self.progress_var.set("")

    def _run(self):
        path = self.file_path.get()
        if path == "No file selected" or not os.path.isfile(path):
            messagebox.showerror("No File", "Please select a valid Excel file.")
            return
        ind   = {k: v.get() for k, v in self.ind_caps.items()}
        theme = core.COLOR_THEMES[self.theme_var.get()]
        self.gen_btn.config(state="disabled", text="⏳ Generating…")
        self.prog_bar.start(10)

        def task():
            try:
                out, warns = generate(path, ind, theme,
                    progress_cb=lambda m: self.after(0, lambda msg=m: self.progress_var.set(msg)))
                self.after(0, lambda: self._ok(out, warns))
            except Exception as e:
                self.after(0, lambda err=str(e): self._err(err))

        threading.Thread(target=task, daemon=True).start()

    def _ok(self, out, warns):
        self.prog_bar.stop()
        self.gen_btn.config(state="normal", text="⚡ Generate Coating Labs Dashboard")
        self.progress_var.set("✅ Done!")
        msg = f"✅ Dashboard created!\n\n📁 {out}\n\n"
        msg += "9 sheets: Weekly Planner, Individual & Combined Utilization,\n"
        msg += "Summary, Charts (4), YoY, Gantt"
        if warns: msg += "\n\n⚠️ " + "\n".join(warns)
        messagebox.showinfo("Success", msg)

    def _err(self, e):
        self.prog_bar.stop()
        self.gen_btn.config(state="normal", text="⚡ Generate Coating Labs Dashboard")
        self.progress_var.set("❌ Error")
        messagebox.showerror("Error", e)


if __name__ == "__main__":
    App().mainloop()
